"""银行流水和银行日记账的自包含输入预检查。"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from statistics import pstdev
from typing import Any, Iterable, Sequence

import pandas as pd


SUPPORTED_SUFFIXES = {".xlsx", ".xls", ".csv"}
SUMMARY_ROW_KEYWORDS = (
    "本日合计", "本日累计", "本日发生额", "本日余额", "本日结存", "本日小计",
    "本旬合计", "本旬累计", "本旬发生额", "本旬余额", "本旬结存", "本旬小计",
    "本月合计", "本月累计", "本月发生额", "本月余额", "本月结存", "本月小计",
    "本季合计", "本季累计", "本季发生额", "本季余额", "本季结存", "本季小计",
    "本年合计", "本年累计", "本年发生额", "本年余额", "本年结存", "本年小计",
    "本期合计", "本期累计", "本期发生额", "本期余额", "本期结存", "本期小计",
    "日计", "月计", "季计", "年计", "期计", "日结", "月结", "季结", "年结",
    "合计", "累计", "总计", "小计", "大计", "发生额", "余额", "结存",
    "本页合计", "本页累计", "本页小计", "过次页", "承前页",
    "期初余额", "期末余额", "期初结存", "期末结存",
    "年初余额", "年末余额", "年初结存", "年末结存",
    "月初余额", "月末余额", "月初结存", "月末结存",
    "结转下年", "结转下期", "结转下月", "上年结转", "上期结转", "上月结转",
    "上年结余", "上期结余", "承前余额", "结转余额",
    "当前合计", "当前累计", "当前余额",
)
HEADER_KEYWORD_GROUPS = (
    ("日期", "date", "交易时间", "记账时间"),
    ("金额", "amount", "发生额"),
    ("借方", "借", "debit", "支出"),
    ("贷方", "贷", "credit", "收入"),
    ("方向", "借贷标志"),
    ("摘要", "summary", "用途", "业务说明"),
    ("余额", "balance"),
    ("对方户名", "对方名称", "附言", "备注"),
    ("凭证", "voucher"),
)


@dataclass(frozen=True)
class HeaderCandidate:
    """一个可追溯的表头范围候选。"""

    skiprows: int
    header_rows: int
    score: float
    columns: tuple[str, ...]
    evidence: tuple[str, ...] = ()


@dataclass(frozen=True)
class TableStructure:
    """实际采用的表头结构和备选证据。"""

    skiprows: int
    header_rows: int
    columns: list[str]
    score: float
    ambiguous: bool = False
    candidates: tuple[HeaderCandidate, ...] = field(default_factory=tuple)
    explanation: str = ""


@dataclass(frozen=True)
class PrecheckItem:
    """一项可展示、可留痕的输入检查结果。"""

    name: str
    bank_result: str
    journal_result: str
    comparison: str
    status: str
    explanation: str

    def as_dict(self) -> dict[str, str]:
        return {
            "检查项目": self.name,
            "银行流水结果": self.bank_result,
            "银行日记账结果": self.journal_result,
            "双方比较结果": self.comparison,
            "状态": self.status,
            "说明": self.explanation,
        }


@dataclass(frozen=True)
class InputPrecheckReport:
    """本次核对的全部输入检查结果。"""

    items: tuple[PrecheckItem, ...]

    @property
    def has_blockers(self) -> bool:
        return any(item.status == "阻止" for item in self.items)

    @property
    def has_warnings(self) -> bool:
        return any(item.status == "提示" for item in self.items)

    def blocker_message(self) -> str:
        return "\n".join(
            f"• {item.name}：{item.explanation}"
            for item in self.items
            if item.status == "阻止"
        )

    def warning_message(self) -> str:
        return "\n".join(
            f"• {item.name}：{item.explanation}"
            for item in self.items
            if item.status == "提示"
        )

    def to_dataframe(self) -> pd.DataFrame:
        return pd.DataFrame(
            [item.as_dict() for item in self.items],
            columns=(
                "检查项目",
                "银行流水结果",
                "银行日记账结果",
                "双方比较结果",
                "状态",
                "说明",
            ),
        )


class InputPrecheckBlockedError(ValueError):
    """输入存在硬错误，正式匹配不得开始。"""

    def __init__(self, report: InputPrecheckReport):
        self.report = report
        super().__init__(report.blocker_message() or "输入预检查未通过")


def _cell_text(value: Any) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() in {"nan", "none"} else text


def _deduplicate(names: Sequence[str]) -> list[str]:
    counts: dict[str, int] = {}
    result = []
    for name in names:
        counts[name] = counts.get(name, 0) + 1
        result.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    return result


def flatten_header_rows(rows: Sequence[Sequence[Any]]) -> list[str]:
    """按列扁平化一至三行表头，保留原始文字顺序。"""
    if not rows:
        return []
    width = max(len(row) for row in rows)
    names = []
    for column_index in range(width):
        parts = []
        for row in rows:
            text = _cell_text(row[column_index] if column_index < len(row) else None)
            if text and (not parts or parts[-1] != text):
                parts.append(text)
        names.append("｜".join(parts) if parts else f"第{column_index + 1}列")
    return _deduplicate(names)


def _read_csv(path: Path, **kwargs: Any) -> pd.DataFrame:
    last_error: Exception | None = None
    for encoding in ("utf-8-sig", "gbk", "latin1"):
        try:
            return pd.read_csv(path, encoding=encoding, **kwargs)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"CSV 文件无法解码：{path.name}") from last_error


def _read_preview(path: Path, max_scan_rows: int) -> tuple[pd.DataFrame, list[tuple[int, int, int, int]]]:
    if path.suffix.lower() == ".csv":
        return (
            _read_csv(path, header=None, nrows=max_scan_rows, dtype=object),
            [],
        )

    preview = pd.read_excel(
        path,
        header=None,
        nrows=max_scan_rows,
        dtype=object,
    )
    merges: list[tuple[int, int, int, int]] = []
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=False, data_only=True)
        sheet = workbook.worksheets[0]
        merges = [
            (
                merged.min_row - 1,
                merged.max_row - 1,
                merged.min_col - 1,
                merged.max_col - 1,
            )
            for merged in sheet.merged_cells.ranges
            if merged.min_row <= max_scan_rows
        ]
        workbook.close()
    return preview, merges


def _expand_merges(
    preview: pd.DataFrame,
    merges: Iterable[tuple[int, int, int, int]],
) -> pd.DataFrame:
    expanded = preview.copy()
    for min_row, max_row, min_col, max_col in merges:
        if min_row >= len(expanded) or min_col >= len(expanded.columns):
            continue
        value = expanded.iat[min_row, min_col]
        for row in range(min_row, min(max_row + 1, len(expanded))):
            for column in range(min_col, min(max_col + 1, len(expanded.columns))):
                expanded.iat[row, column] = value
    return expanded


def _keyword_hits(values: Iterable[Any]) -> int:
    text = " ".join(_cell_text(value).lower() for value in values)
    return sum(
        1
        for group in HEADER_KEYWORD_GROUPS
        if any(keyword.lower() in text for keyword in group)
    )


def _looks_numeric_or_date(value: Any) -> bool:
    if pd.isna(value) or isinstance(value, (datetime, date, int, float)):
        return not pd.isna(value)
    text = str(value).strip().replace(",", "")
    if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", text):
        return True
    return bool(
        re.fullmatch(
            r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}|\d{8}",
            text,
        )
    )


def _candidate_score(
    raw_preview: pd.DataFrame,
    expanded: pd.DataFrame,
    start: int,
    depth: int,
) -> HeaderCandidate | None:
    data_start = start + depth
    if data_start >= len(expanded):
        return None
    header = expanded.iloc[start:data_start]
    raw_header = raw_preview.iloc[start:data_start]
    columns = flatten_header_rows(header.values.tolist())
    if not columns:
        return None

    data = expanded.iloc[data_start:data_start + 5]
    data = data.loc[data.notna().any(axis=1)]
    if data.empty:
        return None

    width = len(expanded.columns)
    densities = [row.notna().sum() / width for _, row in data.iterrows()]
    coverage = sum(name and not name.startswith("第") for name in columns) / width
    keywords = _keyword_hits(header.to_numpy().ravel())
    transitions = 0
    for column_index in range(width):
        header_has_text = any(
            _cell_text(value) and not _looks_numeric_or_date(value)
            for value in header.iloc[:, column_index]
        )
        first_data_value = next(
            (
                value
                for value in data.iloc[:, column_index]
                if _cell_text(value)
            ),
            None,
        )
        if header_has_text and first_data_value is not None and _looks_numeric_or_date(first_data_value):
            transitions += 1

    raw_nonempty = raw_header.notna().sum(axis=1)
    title_penalty = 1 if int(raw_nonempty.iloc[0]) <= 2 else 0
    next_header_penalty = 1 if _keyword_hits(data.iloc[0].tolist()) >= 2 else 0
    data_like_header_ratio = sum(
        _looks_numeric_or_date(value)
        for value in header.iloc[-1]
        if _cell_text(value)
    ) / width
    stability = 1 - min(pstdev(densities), 1) if len(densities) > 1 else 1
    score = (
        keywords * 1.5
        + coverage * 2
        + (transitions / width) * 2
        + sum(densities) / len(densities)
        + stability
        - title_penalty * 4
        - next_header_penalty * 2
        - data_like_header_ratio * 5
        - (depth - 1) * 0.1
    )
    evidence = (
        f"业务词组{keywords}个",
        f"列覆盖率{coverage:.0%}",
        f"数据区稳定度{stability:.0%}",
        f"类型转换列{transitions}个",
        f"表头数据型占比{data_like_header_ratio:.0%}",
    )
    return HeaderCandidate(start, depth, round(score, 4), tuple(columns), evidence)


def detect_table_structure(
    file_path: str | Path,
    max_scan_rows: int = 40,
) -> TableStructure:
    """基于多类结构证据识别银行明细表头。"""
    path = Path(file_path)
    if not path.exists() or not path.is_file():
        raise FileNotFoundError(f"文件不存在：{path}")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise ValueError(f"不支持的文件格式：{path.suffix or '无后缀'}")

    preview, merges = _read_preview(path, max_scan_rows)
    preview = preview.dropna(axis=1, how="all")
    if preview.empty or len(preview.columns) == 0:
        raise ValueError(f"文件没有可识别的表格内容：{path.name}")
    expanded = _expand_merges(preview, merges)

    candidates = []
    max_start = min(len(preview) - 1, max_scan_rows - 1)
    for start in range(max_start):
        for depth in range(1, min(3, len(preview) - start - 1) + 1):
            candidate = _candidate_score(preview, expanded, start, depth)
            if candidate is not None:
                candidates.append(candidate)
    if not candidates:
        raise ValueError(f"无法识别表头和数据区域：{path.name}")

    candidates.sort(key=lambda item: (-item.score, item.skiprows, item.header_rows))
    best = candidates[0]
    runner_up = next(
        (
            item
            for item in candidates[1:]
            if (item.skiprows, item.header_rows) != (best.skiprows, best.header_rows)
        ),
        None,
    )
    ambiguous = bool(runner_up and best.score - runner_up.score <= 0.35)
    explanation = (
        f"首选第{best.skiprows + 1}行起、共{best.header_rows}行表头"
        + ("；存在分数接近的备选结构" if ambiguous else "")
    )
    return TableStructure(
        skiprows=best.skiprows,
        header_rows=best.header_rows,
        columns=list(best.columns),
        score=best.score,
        ambiguous=ambiguous,
        candidates=tuple(candidates[:5]),
        explanation=explanation,
    )


def derive_header_columns(
    file_path: str | Path,
    skiprows: int,
    header_rows: int,
) -> list[str]:
    """按用户采用的表头范围派生稳定列名。"""
    if skiprows < 0 or header_rows < 1:
        raise ValueError("跳过行数不得小于0，表头行数必须大于等于1")
    path = Path(file_path)
    preview, merges = _read_preview(path, skiprows + header_rows)
    preview = preview.dropna(axis=1, how="all")
    if len(preview) < skiprows + header_rows:
        raise ValueError("指定的表头范围超出文件内容")
    expanded = _expand_merges(preview, merges)
    return flatten_header_rows(
        expanded.iloc[skiprows:skiprows + header_rows].values.tolist()
    )


def _required_columns(mapping: dict[str, Any]) -> list[tuple[str, str]]:
    required = [("date", "日期")]
    mode = mapping.get("mode", "debit_credit")
    if mode == "debit_credit":
        required.extend((("debit", "借方/支出"), ("credit", "贷方/收入")))
    elif mode == "single_amount_with_direction":
        required.extend((("amount", "金额"), ("direction", "方向")))
    elif mode == "signed_amount":
        required.append(("amount", "金额"))
    else:
        required.append(("mode", "金额模式"))
    return required


def _mapping_problems(mapping: dict[str, Any], columns: Sequence[Any]) -> list[str]:
    available = set(columns)
    selected = []
    problems = []
    for key, label in _required_columns(mapping):
        value = mapping.get(key)
        if key == "mode" or value in (None, "", "(无)"):
            problems.append(f"缺少{label}列")
        elif value not in available:
            problems.append(f"{label}列不存在：{value}")
        elif value in selected:
            problems.append(f"{label}列与其他必填映射重复：{value}")
        else:
            selected.append(value)
    return problems


def _structure_result(structure: TableStructure) -> str:
    return (
        f"第{structure.skiprows + 1}行起，{structure.header_rows}行表头；"
        f"派生列名：{'、'.join(structure.columns)}"
    )


def _ambiguity_changes_mapping(
    structure: TableStructure,
    mapping: dict[str, Any],
) -> bool:
    if not structure.ambiguous:
        return False
    required_names = {
        mapping.get(key)
        for key, _label in _required_columns(mapping)
        if key != "mode" and mapping.get(key)
    }
    alternatives = [
        candidate
        for candidate in structure.candidates
        if (candidate.skiprows, candidate.header_rows)
        != (structure.skiprows, structure.header_rows)
    ]
    return not alternatives or any(
        not required_names.issubset(set(candidate.columns))
        for candidate in alternatives[:2]
    )


def _valid_dates(frame: pd.DataFrame) -> pd.Series:
    if frame.empty or "date" not in frame.columns:
        return pd.Series(dtype="datetime64[ns]")
    return pd.to_datetime(frame["date"], errors="coerce").dropna()


def _amount_totals(frame: pd.DataFrame) -> tuple[Decimal, Decimal] | None:
    if frame.empty or "amount" not in frame.columns:
        return None
    amounts = [
        Decimal(str(value))
        for value in frame["amount"]
        if pd.notna(value)
    ]
    if not amounts:
        return None
    income = sum((value for value in amounts if value > 0), Decimal("0"))
    expense = sum((-value for value in amounts if value < 0), Decimal("0"))
    return income, expense


def _error_count(
    parse_errors: Sequence[dict[str, Any]],
    source: str,
    error_type: str,
) -> int:
    return sum(
        1
        for error in parse_errors
        if error.get("source_type") == source and error.get("type") == error_type
    )


NON_TRANSACTION_KEYWORDS = (
    "合计",
    "累计",
    "小计",
    "总计",
    "日计",
    "月计",
    "年计",
    "期初",
    "期末",
    "承前页",
    "过次页",
    "统计",
)
NOTE_PREFIXES = ("注：", "说明：", "备注：", "单位：", "制表：")


def _non_transaction_mask(
    frame: pd.DataFrame,
    mapping: dict[str, Any],
) -> pd.Series:
    if frame.empty:
        return pd.Series(False, index=frame.index)
    empty = frame.map(lambda value: not _cell_text(value)).all(axis=1)
    combined = frame.apply(
        lambda row: " ".join(_cell_text(value) for value in row),
        axis=1,
    )
    generic_summary_or_note = combined.str.contains(
        "|".join(re.escape(word) for word in NON_TRANSACTION_KEYWORDS),
        na=False,
    ) | combined.str.startswith(NOTE_PREFIXES)
    column_names = {str(column).strip() for column in frame.columns}
    repeated_header = frame.apply(
        lambda row: sum(
            _cell_text(value) in column_names
            for value in row
            if _cell_text(value)
        )
        >= min(2, len(column_names)),
        axis=1,
    )
    date_column = mapping.get("date")
    dates = (
        pd.to_datetime(frame[date_column], errors="coerce")
        if date_column in frame.columns
        else pd.Series(pd.NaT, index=frame.index)
    )
    summary_column = mapping.get("summary")
    explicit_summary = (
        frame[summary_column].map(_cell_text).isin(SUMMARY_ROW_KEYWORDS)
        if summary_column in frame.columns
        else pd.Series(False, index=frame.index)
    )
    summary = explicit_summary | (generic_summary_or_note & dates.isna())
    nonempty_counts = frame.map(lambda value: bool(_cell_text(value))).sum(axis=1)
    title_or_note = (nonempty_counts == 1) & dates.isna()
    return empty | summary | repeated_header | title_or_note


def _auxiliary_result(
    frame: pd.DataFrame,
    mapping: dict[str, Any],
) -> tuple[str, bool]:
    selected = []
    for column in [mapping.get("summary"), *(mapping.get("auxiliary_text_columns") or [])]:
        if column and column not in selected:
            selected.append(column)
    usable = [column for column in selected if column in frame.columns]
    if not usable:
        return "未选择可用辅助文字列", True
    transaction_rows = frame.loc[~_non_transaction_mask(frame, mapping)]
    denominator = len(transaction_rows)
    if denominator == 0:
        return "没有可计算非空率的交易行", True
    parts = []
    low = False
    for column in usable:
        nonempty = transaction_rows[column].map(lambda value: bool(_cell_text(value))).sum()
        rate = nonempty / denominator
        parts.append(f"{column} {rate:.1%}")
        low = low or rate < 0.8
    missing = [column for column in selected if column not in frame.columns]
    if missing:
        parts.append(f"未找到：{'、'.join(missing)}")
        low = True
    return "；".join(parts), low


def build_input_precheck(
    *,
    raw_bank: pd.DataFrame,
    raw_journal: pd.DataFrame,
    bank: pd.DataFrame,
    journal: pd.DataFrame,
    bank_mapping: dict[str, Any],
    journal_mapping: dict[str, Any],
    bank_structure: TableStructure,
    journal_structure: TableStructure,
    parse_errors: Sequence[dict[str, Any]] = (),
) -> InputPrecheckReport:
    """对已读取和标准化的双方数据执行八项统一检查。"""
    items = []

    file_blocked = raw_bank.empty or raw_journal.empty
    items.append(
        PrecheckItem(
            "文件读取",
            f"可读取，{len(raw_bank)}行" if not raw_bank.empty else "没有可核对数据",
            f"可读取，{len(raw_journal)}行" if not raw_journal.empty else "没有可核对数据",
            "双方均已读取" if not file_blocked else "至少一侧没有数据",
            "阻止" if file_blocked else "通过",
            "文件没有可核对数据，请检查表头和数据区域。" if file_blocked else "文件可正常读取。",
        )
    )

    bank_major = _ambiguity_changes_mapping(bank_structure, bank_mapping)
    journal_major = _ambiguity_changes_mapping(journal_structure, journal_mapping)
    ambiguous = bank_structure.ambiguous or journal_structure.ambiguous
    user_override = any(
        structure.explanation.startswith("采用用户设置")
        for structure in (bank_structure, journal_structure)
    )
    structure_status = (
        "阻止"
        if bank_major or journal_major
        else ("提示" if ambiguous or user_override else "通过")
    )
    items.append(
        PrecheckItem(
            "表格结构",
            _structure_result(bank_structure),
            _structure_result(journal_structure),
            (
                "存在重大歧义"
                if bank_major or journal_major
                else (
                    "采用用户设置"
                    if user_override
                    else ("存在备选结构" if ambiguous else "结构明确")
                )
            ),
            structure_status,
            (
                "表头候选会改变必填列映射，请返回确认表头位置和层级。"
                if structure_status == "阻止"
                else (
                    "当前采用用户设置的表头范围，与程序首选候选不同；列映射仍然有效。"
                    if user_override
                    else ("检测到分数接近的表头候选，请确认当前列映射。" if ambiguous else "表头和数据区域可用。")
                )
            ),
        )
    )

    bank_dates = _valid_dates(bank)
    journal_dates = _valid_dates(journal)
    date_blocked = bank_dates.empty or journal_dates.empty
    bank_date_errors = sum(
        _error_count(parse_errors, "bank", error_type)
        for error_type in ("日期解析失败", "空日期行")
    )
    journal_date_errors = sum(
        _error_count(parse_errors, "journal", error_type)
        for error_type in ("日期解析失败", "空日期行")
    )
    if date_blocked:
        date_status = "阻止"
        date_explanation = "至少一侧日期全部无法解析，请检查日期列和日期格式。"
        date_comparison = "无法比较"
    else:
        bank_range = (bank_dates.min(), bank_dates.max())
        journal_range = (journal_dates.min(), journal_dates.max())
        mismatch = bank_range != journal_range
        date_status = "提示" if mismatch or bank_date_errors or journal_date_errors else "通过"
        date_comparison = "范围一致" if not mismatch else "范围不完全一致"
        date_explanation = (
            f"少量日期或金额解析失败：银行流水日期{bank_date_errors}行，"
            f"银行日记账日期{journal_date_errors}行。"
            if bank_date_errors or journal_date_errors
            else ("双方日期范围不完全一致，请确认是否属于正常未达期间。" if mismatch else "双方日期范围一致。")
        )
    items.append(
        PrecheckItem(
            "日期范围",
            "无法形成日期范围" if bank_dates.empty else f"{bank_dates.min():%Y-%m-%d} 至 {bank_dates.max():%Y-%m-%d}",
            "无法形成日期范围" if journal_dates.empty else f"{journal_dates.min():%Y-%m-%d} 至 {journal_dates.max():%Y-%m-%d}",
            date_comparison,
            date_status,
            date_explanation,
        )
    )

    bank_direction_errors = _error_count(parse_errors, "bank", "方向解析失败")
    journal_direction_errors = _error_count(parse_errors, "journal", "方向解析失败")
    direction_blocked = bool(bank_direction_errors or journal_direction_errors)
    items.append(
        PrecheckItem(
            "金额方向",
            f"银行口径；无法识别{bank_direction_errors}行",
            f"日记账口径；无法识别{journal_direction_errors}行",
            "贷增借减 / 借增贷减",
            "阻止" if direction_blocked else "通过",
            (
                "方向列存在无法识别的值，收入和支出方向不可靠。"
                if direction_blocked
                else "银行流水按贷增借减，银行日记账按借增贷减。"
            ),
        )
    )

    bank_totals = _amount_totals(bank)
    journal_totals = _amount_totals(journal)
    amount_blocked = bank_totals is None or journal_totals is None
    bank_amount_errors = _error_count(parse_errors, "bank", "金额解析失败")
    journal_amount_errors = _error_count(parse_errors, "journal", "金额解析失败")
    if amount_blocked:
        amount_status = "阻止"
        amount_comparison = "无法比较"
        amount_explanation = "至少一侧金额全部无法解析，请检查金额列或借贷列。"
    else:
        income_diff = bank_totals[0] - journal_totals[0]
        expense_diff = bank_totals[1] - journal_totals[1]
        has_diff = income_diff != 0 or expense_diff != 0
        amount_status = "提示" if has_diff or bank_amount_errors or journal_amount_errors else "通过"
        amount_comparison = f"收入差额 {income_diff:.2f}；支出差额 {expense_diff:.2f}"
        amount_explanation = (
            f"少量日期或金额解析失败：银行流水金额{bank_amount_errors}行，"
            f"银行日记账金额{journal_amount_errors}行。"
            if bank_amount_errors or journal_amount_errors
            else ("双方收入或支出合计存在差额；差额是核对对象，不直接视为输入错误。" if has_diff else "双方收入和支出合计一致。")
        )
    items.append(
        PrecheckItem(
            "金额合计",
            "无法形成金额合计" if bank_totals is None else f"收入 {bank_totals[0]:.2f}；支出 {bank_totals[1]:.2f}",
            "无法形成金额合计" if journal_totals is None else f"收入 {journal_totals[0]:.2f}；支出 {journal_totals[1]:.2f}",
            amount_comparison,
            amount_status,
            amount_explanation,
        )
    )

    bank_non_transactions = int(_non_transaction_mask(raw_bank, bank_mapping).sum())
    journal_non_transactions = int(_non_transaction_mask(raw_journal, journal_mapping).sum())
    has_non_transactions = bank_non_transactions > 0 or journal_non_transactions > 0
    items.append(
        PrecheckItem(
            "非交易行",
            f"识别并排除{bank_non_transactions}行",
            f"识别并排除{journal_non_transactions}行",
            f"合计{bank_non_transactions + journal_non_transactions}行",
            "提示" if has_non_transactions else "通过",
            "检测到合计、累计、统计、标题、重复表头、注释或空行。" if has_non_transactions else "未发现混入数据区的非交易行。",
        )
    )

    bank_mapping_problems = _mapping_problems(bank_mapping, raw_bank.columns)
    journal_mapping_problems = _mapping_problems(journal_mapping, raw_journal.columns)
    mapping_blocked = bool(bank_mapping_problems or journal_mapping_problems)
    items.append(
        PrecheckItem(
            "必填字段",
            "完整" if not bank_mapping_problems else "；".join(bank_mapping_problems),
            "完整" if not journal_mapping_problems else "；".join(journal_mapping_problems),
            "双方完整" if not mapping_blocked else "存在缺失或无效映射",
            "阻止" if mapping_blocked else "通过",
            "请返回选择当前金额模式所需的必填列。" if mapping_blocked else "当前金额模式所需字段完整。",
        )
    )

    bank_aux, bank_aux_low = _auxiliary_result(raw_bank, bank_mapping)
    journal_aux, journal_aux_low = _auxiliary_result(raw_journal, journal_mapping)
    aux_low = bank_aux_low or journal_aux_low
    items.append(
        PrecheckItem(
            "辅助文字完整性",
            bank_aux,
            journal_aux,
            "至少一侧偏低" if aux_low else "双方可用",
            "提示" if aux_low else "通过",
            "摘要、对方户名等辅助文字非空率低于80%，文字匹配证据可能不足。" if aux_low else "辅助文字列可用。",
        )
    )

    return InputPrecheckReport(items=tuple(items))
