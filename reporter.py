"""
Reporter 模块 — 核对结果报表生成器

使用 make_excel deep-navy 主题输出 Excel，再通过 openpyxl 后处理添加条件格式。
"""

from typing import Optional, List, Dict, Any, Callable
from decimal import Decimal
from dataclasses import dataclass
from datetime import datetime
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation

from precision_engine import PrecisionEngine
from data_structures import InitialBalanceWarning, MatcherConfig
from data_loader import ParseErrorCollector
from input_precheck import InputPrecheckReport
from matcher import Matcher
from utils import round_decimal, clean_excel_string
from balance import BalanceRecalculator, BalanceReconciler
from make_excel import make_excel
from llm_assistant import redact_sensitive_text, sanitize_url


def _restore_numeric_cells(value: Any) -> Any:
    """将清洗后的字符串还原为数值（支持千分位文本），无法还原时原样返回。

    未达明细直接展示原始数据行，金额/余额列经 clean_excel_string 后变成字符串，
    在 Excel 中无法求和；此函数把可解析的值还原为 float，文本保持文本。
    """
    if not isinstance(value, str):
        return value
    s = value.strip().replace(',', '')
    if not s or s.lower() == 'nan':
        return value
    try:
        return float(s)
    except ValueError:
        return value


class Reporter:
    """核对结果报表生成器，使用 make_excel deep-navy 主题输出 Excel。"""

    METRICS = ['income_count', 'income_amount', 'expense_count', 'expense_amount']
    SUFFIXES = ['_bank', '_journal']
    RENAME_MAP_DAILY = {
        'date': '日期', 'income_count_bank': '银行-收入笔数', 'income_count_journal': '日记账-收入笔数',
        'income_count_diff': '收入笔数差额', 'income_amount_bank': '银行-收入金额', 'income_amount_journal': '日记账-收入金额',
        'income_amount_diff': '收入金额差额', 'expense_count_bank': '银行-支出笔数', 'expense_count_journal': '日记账-支出笔数',
        'expense_count_diff': '支出笔数差额', 'expense_amount_bank': '银行-支出金额', 'expense_amount_journal': '日记账-支出金额',
        'expense_amount_diff': '支出金额差额', 'net_bank': '银行-变动净额', 'net_journal': '日记账-变动净额',
        'balance_bank': '银行-期末余额', 'balance_journal': '日记账-期末余额', 'balance_diff': '余额差额',
    }
    RENAME_MAP_MONTHLY = {
        'month': '月份', 'income_count_bank': '银行-收入笔数', 'income_count_journal': '日记账-收入笔数',
        'income_count_diff': '收入笔数差额', 'income_amount_bank': '银行-收入金额', 'income_amount_journal': '日记账-收入金额',
        'income_amount_diff': '收入金额差额', 'expense_count_bank': '银行-支出笔数', 'expense_count_journal': '日记账-支出笔数',
        'expense_count_diff': '支出笔数差额', 'expense_amount_bank': '银行-支出金额', 'expense_amount_journal': '日记账-支出金额',
        'expense_amount_diff': '支出金额差额', 'net_bank': '银行-变动净额', 'net_journal': '日记账-变动净额',
        'balance_bank': '银行-月末余额', 'balance_journal': '日记账-月末余额', 'balance_diff': '余额差额',
    }

    def __init__(self, matcher: Matcher, raw_bank: Optional[pd.DataFrame] = None,
                 raw_journal: Optional[pd.DataFrame] = None,
                 bank_mapping: Optional[Dict[str, Any]] = None,
                  journal_mapping: Optional[Dict[str, Any]] = None,
                  logger: Optional[Callable[[str], None]] = None,
                  error_collector: Optional[ParseErrorCollector] = None,
                  precheck_report: Optional[InputPrecheckReport] = None):
        self.matcher = matcher
        self.raw_bank = raw_bank
        self.raw_journal = raw_journal
        self.bank_mapping = bank_mapping
        self.journal_mapping = journal_mapping
        self.logger = logger
        self.initial_balance_warning: Optional[InitialBalanceWarning] = None
        self.error_collector = error_collector
        self.precheck_report = precheck_report

    @staticmethod
    def _get_ordered_columns(df: pd.DataFrame, date_col: str = 'date') -> List[str]:
        """按照指标顺序排列列名。"""
        new_columns = [date_col]
        for metric in Reporter.METRICS:
            for suffix in Reporter.SUFFIXES:
                if metric + suffix in df.columns:
                    new_columns.append(metric + suffix)
            if metric + '_diff' in df.columns:
                new_columns.append(metric + '_diff')
        for suffix in Reporter.SUFFIXES:
            if 'net' + suffix in df.columns:
                new_columns.append('net' + suffix)
        for suffix in Reporter.SUFFIXES:
            if 'balance' + suffix in df.columns:
                new_columns.append('balance' + suffix)
        if 'balance_diff' in df.columns:
            new_columns.append('balance_diff')
        return new_columns

    @staticmethod
    def _has_balance_data(df: Optional[pd.DataFrame], mapping: Optional[Dict[str, Any]] = None) -> bool:
        """判断数据中是否存在可用的余额列数据（mapping 列优先，否则按内置词表精确匹配）。"""
        if df is None or df.empty:
            return False
        col = None
        if mapping:
            cand = mapping.get('balance')
            if cand and cand in df.columns:
                col = cand
        if col is None:
            for c in df.columns:
                c_stripped, c_lower = str(c).strip(), str(c).lower().strip()
                if c_stripped in ('balance', '余额', 'std_balance') or c_lower == 'balance':
                    col = c
                    break
        if col is None:
            return False
        series = df[col]
        non_empty = (series.notna()
                     & (series.astype(str).str.strip() != '')
                     & (series.astype(str).str.strip().str.lower() != 'nan'))
        return bool(non_empty.any())

    def check_balance_continuity(self, df: pd.DataFrame, tolerance_li: int = 10,
                                 source: str = "") -> List[Dict[str, Any]]:
        """按日检查余额连续性：以最近一个有效余额日为基准，用累计净额推算预期余额。

        空余额日安全跳过且不断链——其净额计入累计，下一个有效余额日仍可校验。
        """
        anomalies: List[Dict[str, Any]] = []
        if df.empty or 'balance' not in df.columns or 'amount' not in df.columns:
            return anomalies
        sorted_df = df.sort_values(['date', 'original_idx'])
        daily_end = sorted_df.groupby('date').last()
        daily_net = df.groupby('date')['amount'].sum().reset_index()
        merged = pd.merge(daily_end[['balance']], daily_net, left_index=True, right_on='date', how='left')

        tolerance_yuan = PrecisionEngine.from_integer_li(tolerance_li)
        prev_valid_balance = None
        cum_net = Decimal('0')  # 自最近一个有效余额日之后（不含当日）的累计净额
        for _, row in merged.iterrows():
            balance, net = row['balance'], row['amount']
            net_val = Decimal('0') if pd.isna(net) else net
            if pd.isna(balance):
                # 空余额日：无法校验，但当日净额计入累计，不断链
                cum_net += net_val
                continue
            if prev_valid_balance is not None:
                period_net = cum_net + net_val
                expected = prev_valid_balance + period_net
                diff = abs(balance - expected)
                if Decimal(str(diff)) > tolerance_yuan:
                    anomalies.append({
                        '来源': source, '日期': row['date'], '基准余额': prev_valid_balance,
                        '区间净额': period_net, '预期余额': expected,
                        '实际余额': balance, '差额': diff
                    })
            prev_valid_balance = balance
            cum_net = Decimal('0')
        return anomalies

    def calculate_daily_stats(self, bank: pd.DataFrame, journal: pd.DataFrame) -> pd.DataFrame:
        """计算每日统计对比数据。"""
        if bank.empty and journal.empty:
            return pd.DataFrame()
        all_dates = set(bank['date'].unique()) | set(journal['date'].unique())
        if not all_dates:
            return pd.DataFrame()

        date_range = pd.date_range(start=min(all_dates), end=max(all_dates), freq='D')

        def get_stats(df, has_bal):
            inc = df[df['amount'] > 0]
            exp = df[df['amount'] < 0]
            stats = pd.merge(
                inc.groupby('date')['amount'].agg(income_count='count', income_amount='sum'),
                exp.groupby('date')['amount'].agg(expense_count='count', expense_amount='sum'),
                on='date', how='outer'
            )
            stats = pd.merge(stats, df.groupby('date')['amount'].agg(net='sum'), on='date', how='outer')
            stats = pd.merge(pd.DataFrame({'date': date_range}), stats, on='date', how='left').fillna(0)
            if has_bal:
                bal = df.sort_values(['date', 'original_idx']).groupby('date')['balance'].last()
                stats = pd.merge(stats, bal, on='date', how='left')
                stats['balance'] = stats['balance'].ffill()
                stats['balance_missing'] = stats['balance'].isna()
            return stats

        b_stats = get_stats(bank, 'balance' in bank.columns and bank['balance'].notna().any())
        j_stats = get_stats(journal, 'balance' in journal.columns and journal['balance'].notna().any())

        df = pd.merge(b_stats, j_stats, on='date', how='outer', suffixes=('_bank', '_journal'))

        # 四舍五入与差额计算
        cols = [c for c in df.columns if 'amount' in c or 'net' in c or ('balance' in c and 'missing' not in c)]
        for c in cols:
            df[c] = df[c].apply(round_decimal)

        df['income_count_diff'] = df['income_count_bank'] - df['income_count_journal']
        df['income_amount_diff'] = (df['income_amount_bank'] - df['income_amount_journal']).apply(round_decimal)
        df['expense_count_diff'] = df['expense_count_bank'] - df['expense_count_journal']
        df['expense_amount_diff'] = (df['expense_amount_bank'] - df['expense_amount_journal']).apply(round_decimal)

        if 'balance_bank' in df.columns and 'balance_journal' in df.columns:
            df['balance_diff'] = (df['balance_bank'] - df['balance_journal']).apply(round_decimal)

        return df

    def calculate_monthly_stats(self, df_daily: pd.DataFrame) -> pd.DataFrame:
        """根据每日统计计算月度汇总。"""
        df = df_daily.copy()
        df['month'] = df['date'].dt.to_period('M')

        agg = {k: 'sum' for k in [
            'income_count_bank', 'income_amount_bank', 'expense_count_bank', 'expense_amount_bank', 'net_bank',
            'income_count_journal', 'income_amount_journal', 'expense_count_journal', 'expense_amount_journal', 'net_journal'
        ]}
        df_m = df.groupby('month').agg(agg).reset_index()

        last = df.sort_values('date').drop_duplicates('month', keep='last')
        bal_cols = [c for c in ['balance_bank', 'balance_journal'] if c in last.columns]
        if bal_cols:
            df_m = pd.merge(df_m, last[['month'] + bal_cols], on='month', how='left')
            for c in bal_cols:
                df_m[c] = df_m[c].ffill()

        cols = [c for c in df_m.columns if 'amount' in c or 'net' in c or 'balance' in c]
        for c in cols:
            df_m[c] = df_m[c].apply(round_decimal)

        df_m['income_count_diff'] = df_m['income_count_bank'] - df_m['income_count_journal']
        df_m['income_amount_diff'] = (df_m['income_amount_bank'] - df_m['income_amount_journal']).apply(round_decimal)
        df_m['expense_count_diff'] = df_m['expense_count_bank'] - df_m['expense_count_journal']
        df_m['expense_amount_diff'] = (df_m['expense_amount_bank'] - df_m['expense_amount_journal']).apply(round_decimal)
        if 'balance_bank' in df_m.columns and 'balance_journal' in df_m.columns:
            df_m['balance_diff'] = (df_m['balance_bank'] - df_m['balance_journal']).apply(round_decimal)

        df_m = df_m[self._get_ordered_columns(df_m, 'month')]
        # Period 类型转为字符串
        df_m['month'] = df_m['month'].astype(str)
        return df_m

    # ------------------------------------------------------------------
    # 后处理辅助方法
    # ------------------------------------------------------------------

    @staticmethod
    def _postprocess_summary(ws, has_initial_warning: bool) -> None:
        """汇总表后处理：期初警告行添加红字/黄底。"""
        if not has_initial_warning:
            return
        warning_font = Font(color='FF0000', bold=True, size=14)
        warning_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # 警告块为前 6 行数据（第 1 行是表头），即数据行 2~7
        max_row = ws.max_row or 7
        max_col = ws.max_column or 3
        for r in range(2, min(max_row + 1, 8)):
            for c in range(2, max_col + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = warning_font
                cell.fill = warning_fill

    @staticmethod
    def _postprocess_details(ws) -> None:
        """匹配明细后处理：低置信度行黄底红字，差额列条件格式。"""
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1

        # 查找关键列索引
        low_conf_col_idx = None
        diff_col_indices = []
        amount_keywords = ['金额', '余额', '净额', '差额', '收入', '支出']

        for ci in range(2, max_col + 1):
            header = ws.cell(row=1, column=ci).value
            if header is None:
                continue
            header_str = str(header)
            header_lower = header_str.lower()
            if header_str == '_低置信度标记':
                low_conf_col_idx = ci
            if '差额' in header_str or 'diff' in header_lower:
                diff_col_indices.append(ci)

        low_conf_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        low_conf_font = Font(color='FF0000', bold=True)
        positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for r in range(2, max_row + 1):
            # 低置信度行着色
            is_low_conf = False
            if low_conf_col_idx:
                val = ws.cell(row=r, column=low_conf_col_idx).value
                if val is True or str(val).lower() == 'true':
                    is_low_conf = True
                    for c in range(2, max_col + 1):
                        ws.cell(row=r, column=c).fill = low_conf_fill
                        ws.cell(row=r, column=c).font = low_conf_font

            # 差额列条件格式
            for dc in diff_col_indices:
                cell = ws.cell(row=r, column=dc)
                try:
                    v = float(cell.value) if cell.value is not None else 0
                    if v > 0:
                        cell.fill = positive_fill
                    elif v < 0:
                        cell.fill = negative_fill
                except (ValueError, TypeError):
                    pass

        # 删除 _低置信度标记 列
        if low_conf_col_idx:
            ws.delete_cols(low_conf_col_idx)

    @staticmethod
    def _postprocess_diff_columns(ws) -> None:
        """对含差额列的工作表应用正绿负红条件格式。"""
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        positive_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        negative_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

        for ci in range(2, max_col + 1):
            header = ws.cell(row=1, column=ci).value
            if header is None:
                continue
            header_str = str(header)
            if '差额' in header_str or 'diff' in header_str.lower():
                for r in range(2, max_row + 1):
                    cell = ws.cell(row=r, column=ci)
                    try:
                        v = float(cell.value) if cell.value is not None else 0
                        if v > 0:
                            cell.fill = positive_fill
                        elif v < 0:
                            cell.fill = negative_fill
                    except (ValueError, TypeError):
                        pass

    # ------------------------------------------------------------------
    # 结构化候选报告（新版本）
    # ------------------------------------------------------------------

    @staticmethod
    def _numeric(value: Any) -> Any:
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            return value
        return value

    @staticmethod
    def _safe_table(frame: pd.DataFrame) -> pd.DataFrame:
        """保留数值和日期类型，只清理所有外来文字。"""
        safe = frame.copy()
        for column in safe.columns:
            def clean_value(value):
                if value is None:
                    return ""
                if isinstance(value, Decimal):
                    return float(value)
                if isinstance(value, (int, float, pd.Timestamp, datetime)):
                    return value
                return clean_excel_string(value)
            safe[column] = safe[column].map(clean_value)
        return safe

    def _prepare_initial_balance(self) -> tuple[bool, bool]:
        bank = self.matcher.bank
        journal = self.matcher.journal
        journal_source = (
            self.raw_journal
            if self.raw_journal is not None
            else journal
        )
        journal_initial = BalanceRecalculator.extract_initial_balance(
            journal_source,
            self.journal_mapping,
        )
        bank_initial = BalanceRecalculator.extract_initial_balance(
            bank,
            self.bank_mapping,
        )
        initial_diff = abs(bank_initial - journal_initial)
        journal_has_balance = self._has_balance_data(
            journal_source,
            self.journal_mapping,
        )
        bank_has_balance = self._has_balance_data(
            bank,
            self.bank_mapping,
        )
        balance_check_possible = bank_has_balance and journal_has_balance
        has_warning = (
            balance_check_possible and initial_diff > Decimal("0.01")
        )
        self.initial_balance_warning = InitialBalanceWarning(
            has_warning=has_warning,
            bank_initial=bank_initial,
            journal_initial=journal_initial,
            diff=initial_diff,
            message=(
                "期初余额不一致，请先核对期初余额"
                if has_warning
                else ""
            ),
        )
        return balance_check_possible, has_warning

    @staticmethod
    def _append_source_stats(
        rows: list[tuple[str, Any]],
        frame: pd.DataFrame,
        name: str,
    ) -> None:
        income = frame[frame["amount"] > 0]
        expense = frame[frame["amount"] < 0]
        rows.extend(
            [
                (f"{name}总笔数", int(len(frame))),
                (f"{name}总金额", float(frame["amount"].sum())),
                (f"{name}收入笔数", int(len(income))),
                (f"{name}收入金额", float(income["amount"].sum())),
                (f"{name}支出笔数", int(len(expense))),
                (f"{name}支出金额", float(expense["amount"].sum())),
            ]
        )

    def _build_summary_table(
        self,
        config: MatcherConfig,
        *,
        balance_check_possible: bool,
        has_warning: bool,
    ) -> pd.DataFrame:
        bank = self.matcher.bank
        journal = self.matcher.journal
        candidates = list(
            getattr(self.matcher, "selected_candidates", [])
        )
        rows: list[tuple[str, Any]] = []
        warning = self.initial_balance_warning
        if has_warning and warning is not None:
            rows.extend(
                [
                    ("期初余额差异警告", "请先核对期初余额"),
                    ("银行期初余额", float(warning.bank_initial)),
                    ("日记账期初余额", float(warning.journal_initial)),
                    ("期初余额差额", float(warning.diff)),
                    ("期初余额状态", "不一致"),
                    ("处理提示", "请先核对期初余额"),
                ]
            )

        self._append_source_stats(rows, bank, "银行流水")
        self._append_source_stats(rows, journal, "日记账")

        exact_bank: set[int] = set()
        exact_journal: set[int] = set()
        automatic_bank: set[int] = set()
        automatic_journal: set[int] = set()
        low_confidence_groups = 0
        for candidate in candidates:
            if candidate.metrics.total_diff_li == 0:
                exact_bank.update(candidate.bank_idxs)
                exact_journal.update(candidate.journal_idxs)
            status_value = getattr(
                candidate.processing_status,
                "value",
                str(candidate.processing_status),
            )
            if status_value == "自动确认":
                automatic_bank.update(candidate.bank_idxs)
                automatic_journal.update(candidate.journal_idxs)
            if candidate.scores.total < config.auto_confirm_score:
                low_confidence_groups += 1

        total_rows = len(bank) + len(journal)
        exact_rate = (
            (len(exact_bank) + len(exact_journal)) / total_rows
            if total_rows
            else 0.0
        )
        automatic_rate = (
            (len(automatic_bank) + len(automatic_journal)) / total_rows
            if total_rows
            else 0.0
        )
        low_ratio = (
            low_confidence_groups / len(candidates)
            if candidates
            else 0.0
        )
        automatic_candidates = [
            candidate
            for candidate in candidates
            if candidate.processing_status.value == "自动确认"
        ]
        trivial_candidates = [
            candidate
            for candidate in automatic_candidates
            if candidate.metrics.total_diff_li > 0
        ]
        standalone_pending = [
            candidate
            for candidate in candidates
            if candidate.processing_status.value == "待人工复核"
            and not candidate.evidence.get(
                "included_in_pool_review",
                False,
            )
        ]
        pools = list(
            getattr(self.matcher, "difference_pools", [])
        )
        pending_pools = [
            pool
            for pool in pools
            if pool.exceeds_performance_materiality
        ]
        bank_unmatched = bank[~bank["matched"]]
        journal_unmatched = journal[~journal["matched"]]
        llm_records = list(
            getattr(self.matcher, "llm_records", [])
        )
        assistant_config = getattr(
            getattr(self.matcher, "llm_assistant", None),
            "config",
            None,
        )
        llm_status_changes = sum(
            1
            for candidate in candidates
            if candidate.llm_decision is not None
            and (
                int(
                    candidate.evidence.get(
                        "pre_llm_total_score",
                        candidate.scores.total,
                    )
                )
                < config.auto_confirm_score
                <= candidate.scores.total
            )
        )

        def group_amount(items) -> float:
            return float(
                sum(
                    PrecisionEngine.from_integer_li(
                        item.metrics.group_amount_li
                    )
                    for item in items
                )
            )

        rows.extend(
            [
                ("银行已找到候选笔数", int(bank["matched"].sum())),
                ("日记账已找到候选笔数", int(journal["matched"].sum())),
                (
                    "银行候选覆盖率",
                    float(bank["matched"].mean()) if len(bank) else 0.0,
                ),
                (
                    "日记账候选覆盖率",
                    float(journal["matched"].mean())
                    if len(journal)
                    else 0.0,
                ),
                ("银行自动确认笔数", len(automatic_bank)),
                ("日记账自动确认笔数", len(automatic_journal)),
                (
                    "银行自动确认率",
                    len(automatic_bank) / len(bank) if len(bank) else 0.0,
                ),
                (
                    "日记账自动确认率",
                    len(automatic_journal) / len(journal)
                    if len(journal)
                    else 0.0,
                ),
                ("精确匹配率", float(min(1.0, exact_rate))),
                ("自动处理率", float(min(1.0, automatic_rate))),
                ("匹配组数", int(len(candidates))),
                ("自动确认组数", len(automatic_candidates)),
                ("自动确认金额", group_amount(automatic_candidates)),
                ("明显微小错报组数", len(trivial_candidates)),
                (
                    "明显微小错报金额",
                    float(
                        sum(
                            PrecisionEngine.from_integer_li(
                                candidate.metrics.total_diff_li
                            )
                            for candidate in trivial_candidates
                        )
                    ),
                ),
                ("低可信度组数", int(low_confidence_groups)),
                ("低可信度组占比", float(min(1.0, low_ratio))),
                (
                    "待人工复核事项数",
                    len(standalone_pending) + len(pending_pools),
                ),
                (
                    "待人工复核金额",
                    group_amount(standalone_pending)
                    + float(
                        sum(
                            PrecisionEngine.from_integer_li(
                                pool.total_diff_li
                            )
                            for pool in pending_pools
                        )
                    ),
                ),
                (
                    "银行未找到候选笔数",
                    len(bank_unmatched),
                ),
                (
                    "银行未找到候选金额",
                    float(bank_unmatched["amount"].abs().sum())
                    if not bank_unmatched.empty
                    else 0.0,
                ),
                (
                    "日记账未找到候选笔数",
                    len(journal_unmatched),
                ),
                (
                    "日记账未找到候选金额",
                    float(journal_unmatched["amount"].abs().sum())
                    if not journal_unmatched.empty
                    else 0.0,
                ),
                (
                    "明显微小错报池数",
                    len(pools),
                ),
                (
                    "大模型参与组数",
                    len(llm_records),
                ),
                (
                    "大模型辅助",
                    "启用"
                    if assistant_config is not None
                    and getattr(assistant_config, "enabled", False)
                    else "关闭",
                ),
                (
                    "大模型成功次数",
                    sum(
                        1
                        for record in llm_records
                        if not record.fallback_used
                    ),
                ),
                (
                    "大模型降级次数",
                    sum(
                        1
                        for record in llm_records
                        if record.fallback_used
                    ),
                ),
                ("大模型状态变化数", llm_status_changes),
                ("实际执行重要性水平", float(config.performance_materiality)),
                (
                    "明显微小错报临界值",
                    float(config.clearly_trivial_threshold),
                ),
                ("自动确认最低综合可信度", config.auto_confirm_score),
            ]
        )
        for pool in pools:
            rows.append(
                (
                    (
                        f"差异池累计|{pool.month}|"
                        f"{pool.pool_type.value}"
                    ),
                    float(
                        PrecisionEngine.from_integer_li(
                            pool.total_diff_li
                        )
                    ),
                )
            )

        if warning is not None and not has_warning:
            if balance_check_possible:
                rows.extend(
                    [
                        ("银行期初余额", float(warning.bank_initial)),
                        ("日记账期初余额", float(warning.journal_initial)),
                        ("期初余额差额", float(warning.diff)),
                        ("期初余额状态", "一致"),
                    ]
                )
            else:
                rows.append(
                    (
                        "期初余额核对",
                        "任一方未提供有效余额列数据，跳过期初核对",
                    )
                )
        return pd.DataFrame(rows, columns=["项目", "数值"])

    def _build_daily_and_monthly_tables(
        self,
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        daily = self.calculate_daily_stats(
            self.matcher.bank,
            self.matcher.journal,
        )
        daily_columns = list(self.RENAME_MAP_DAILY.values())
        monthly_columns = list(self.RENAME_MAP_MONTHLY.values())
        if daily.empty:
            return (
                pd.DataFrame(columns=daily_columns),
                pd.DataFrame(columns=monthly_columns),
            )
        ordered = daily[self._get_ordered_columns(daily)]
        daily_cn = ordered.rename(
            columns={
                key: value
                for key, value in self.RENAME_MAP_DAILY.items()
                if key in ordered.columns
            }
        )
        monthly = self.calculate_monthly_stats(daily)
        monthly_cn = monthly.rename(
            columns={
                key: value
                for key, value in self.RENAME_MAP_MONTHLY.items()
                if key in monthly.columns
            }
        )
        return daily_cn, monthly_cn

    @staticmethod
    def _match_type_name(match_type: str) -> str:
        names = {
            "exact_1to1": "精确一对一",
            "tolerance_date": "日期容差",
            "amount_difference": "明显微小金额差异",
            "batch_aggregation": "批量聚合",
            "continuous_summary_group": "连续摘要整组",
            "combination_dfs": "组合求和",
            "daily_total": "日总额",
            "monthly_total": "月总额",
            "cross_month_total": "跨月多对多",
        }
        return names.get(match_type, match_type)

    def _candidate_group_rows(self) -> list[dict[str, Any]]:
        rows: list[dict[str, Any]] = []
        candidates = list(
            getattr(self.matcher, "selected_candidates", [])
        )
        for candidate in candidates:
            all_dates = candidate.bank_dates + candidate.journal_dates
            bank_total_li = (
                candidate.metrics.bank_income_li
                - candidate.metrics.bank_expense_li
            )
            journal_total_li = (
                candidate.metrics.journal_income_li
                - candidate.metrics.journal_expense_li
            )
            evidence = candidate.text_evidence
            difference_pool_ids = "；".join(
                str(value)
                for value in candidate.evidence.get(
                    "difference_pool_ids",
                    [],
                )
            )
            rows.append(
                {
                    "匹配ID": candidate.final_match_id
                    or candidate.candidate_id,
                    "候选ID": candidate.candidate_id,
                    "阶段": candidate.match_stage,
                    "类型": self._match_type_name(candidate.match_type),
                    "最终状态": candidate.processing_status.value,
                    "处理原因": candidate.processing_reason,
                    "综合可信度": candidate.scores.total,
                    "金额分": candidate.scores.amount,
                    "日期分": candidate.scores.date,
                    "文字分": candidate.scores.text,
                    "结构分": candidate.scores.structure,
                    "银行笔数": len(candidate.bank_idxs),
                    "日记账笔数": len(candidate.journal_idxs),
                    "银行合计": float(
                        PrecisionEngine.from_integer_li(bank_total_li)
                    ),
                    "日记账合计": float(
                        PrecisionEngine.from_integer_li(journal_total_li)
                    ),
                    "银行收入": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.bank_income_li
                        )
                    ),
                    "银行支出": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.bank_expense_li
                        )
                    ),
                    "银行净额": float(
                        PrecisionEngine.from_integer_li(bank_total_li)
                    ),
                    "日记账收入": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.journal_income_li
                        )
                    ),
                    "日记账支出": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.journal_expense_li
                        )
                    ),
                    "日记账净额": float(
                        PrecisionEngine.from_integer_li(journal_total_li)
                    ),
                    "组金额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.group_amount_li
                        )
                    ),
                    "收入差额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.income_diff_li
                        )
                    ),
                    "支出差额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.expense_diff_li
                        )
                    ),
                    "总差额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.total_diff_li
                        )
                    ),
                    "银_日期": min(candidate.bank_dates)
                    if candidate.bank_dates
                    else "",
                    "账_日期": min(candidate.journal_dates)
                    if candidate.journal_dates
                    else "",
                    "最早日期": min(all_dates) if all_dates else "",
                    "最晚日期": max(all_dates) if all_dates else "",
                    "文字支持": "；".join(
                        evidence.supporting_fields if evidence else ()
                    ),
                    "文字冲突": "；".join(
                        evidence.conflicting_fields if evidence else ()
                    ),
                    "大模型判断": (
                        candidate.llm_decision.reason
                        if candidate.llm_decision
                        else ""
                    ),
                    "差异池ID": difference_pool_ids,
                    "是否使用大模型": (
                        "是" if candidate.llm_decision else "否"
                    ),
                    "纳入整池复核": (
                        "是"
                        if candidate.evidence.get(
                            "included_in_pool_review",
                            False,
                        )
                        else "否"
                    ),
                }
            )

        if rows:
            return rows

        bank = self.matcher.bank
        journal = self.matcher.journal
        ids = sorted(
            {
                value
                for value in list(bank["match_id"]) + list(journal["match_id"])
                if value
            },
            key=str,
        )
        for match_id in ids:
            bank_group = bank[bank["match_id"] == match_id]
            journal_group = journal[journal["match_id"] == match_id]
            dates = list(bank_group["date"]) + list(journal_group["date"])
            match_type = (
                bank_group.iloc[0]["match_type"]
                if not bank_group.empty
                else journal_group.iloc[0]["match_type"]
            )
            confidence = (
                bank_group.iloc[0]["confidence"]
                if not bank_group.empty
                else journal_group.iloc[0]["confidence"]
            )
            rows.append(
                {
                    "匹配ID": str(match_id),
                    "候选ID": "",
                    "阶段": "",
                    "类型": self._match_type_name(str(match_type)),
                    "最终状态": "自动确认",
                    "处理原因": "兼容旧匹配记录",
                    "综合可信度": {"高": 90, "中": 75, "低": 60}.get(
                        confidence,
                        0,
                    ),
                    "金额分": "",
                    "日期分": "",
                    "文字分": "",
                    "结构分": "",
                    "银行笔数": len(bank_group),
                    "日记账笔数": len(journal_group),
                    "银行合计": float(bank_group["amount"].sum()),
                    "日记账合计": float(journal_group["amount"].sum()),
                    "银行收入": float(
                        bank_group.loc[
                            bank_group["amount"] > 0,
                            "amount",
                        ].sum()
                    ),
                    "银行支出": float(
                        -bank_group.loc[
                            bank_group["amount"] < 0,
                            "amount",
                        ].sum()
                    ),
                    "银行净额": float(bank_group["amount"].sum()),
                    "日记账收入": float(
                        journal_group.loc[
                            journal_group["amount"] > 0,
                            "amount",
                        ].sum()
                    ),
                    "日记账支出": float(
                        -journal_group.loc[
                            journal_group["amount"] < 0,
                            "amount",
                        ].sum()
                    ),
                    "日记账净额": float(journal_group["amount"].sum()),
                    "组金额": max(
                        float(bank_group["amount"].abs().sum()),
                        float(journal_group["amount"].abs().sum()),
                    ),
                    "收入差额": "",
                    "支出差额": "",
                    "总差额": abs(
                        float(bank_group["amount"].sum())
                        - float(journal_group["amount"].sum())
                    ),
                    "银_日期": bank_group["date"].min()
                    if not bank_group.empty
                    else "",
                    "账_日期": journal_group["date"].min()
                    if not journal_group.empty
                    else "",
                    "最早日期": min(dates) if dates else "",
                    "最晚日期": max(dates) if dates else "",
                    "文字支持": "",
                    "文字冲突": "",
                    "大模型判断": "",
                    "差异池ID": "",
                    "是否使用大模型": "否",
                    "纳入整池复核": "否",
                }
            )
        return rows

    def _build_match_group_table(self) -> pd.DataFrame:
        columns = [
            "匹配ID", "候选ID", "阶段", "类型", "最终状态", "处理原因",
            "综合可信度", "金额分", "日期分", "文字分", "结构分",
            "银行笔数", "日记账笔数", "银行合计", "日记账合计",
            "银行收入", "银行支出", "银行净额", "日记账收入",
            "日记账支出", "日记账净额", "组金额", "收入差额",
            "支出差额", "总差额", "银_日期",
            "账_日期", "最早日期", "最晚日期", "文字支持", "文字冲突",
            "大模型判断", "差异池ID", "是否使用大模型",
            "纳入整池复核",
        ]
        rows = self._candidate_group_rows()
        rows.sort(
            key=lambda row: (
                row.get("最早日期") or pd.Timestamp.max,
                row.get("组金额", 0),
                str(row.get("匹配ID", "")),
            )
        )
        return pd.DataFrame(rows, columns=columns)

    @staticmethod
    def _auxiliary_text(row: pd.Series) -> str:
        fields = row.get("aux_text_fields", {})
        if isinstance(fields, dict):
            return json.dumps(fields, ensure_ascii=False, sort_keys=True)
        return ""

    def _build_match_component_table(self) -> pd.DataFrame:
        columns = [
            "匹配ID", "候选ID", "来源", "原文件行号", "日期", "金额",
            "摘要", "辅助文字", "凭证号", "类型", "处理状态",
            "纳入整池复核",
        ]
        rows: list[dict[str, Any]] = []
        candidates = list(
            getattr(self.matcher, "selected_candidates", [])
        )
        if candidates:
            for candidate in sorted(
                candidates,
                key=lambda item: (
                    min(item.bank_dates + item.journal_dates)
                    if item.bank_dates or item.journal_dates
                    else pd.Timestamp.max,
                    item.final_match_id,
                ),
            ):
                for source_name, frame, indexes in (
                    ("银行流水", self.matcher.bank, candidate.bank_idxs),
                    ("日记账", self.matcher.journal, candidate.journal_idxs),
                ):
                    for index in indexes:
                        row = frame.loc[index]
                        rows.append(
                            {
                                "匹配ID": candidate.final_match_id,
                                "候选ID": candidate.candidate_id,
                                "来源": source_name,
                                "原文件行号": int(
                                    row.get(
                                        "original_file_row",
                                        row.get("original_idx", index),
                                    )
                                ),
                                "日期": row.get("date", ""),
                                "金额": float(row.get("amount", 0)),
                                "摘要": row.get("summary", ""),
                                "辅助文字": self._auxiliary_text(row),
                                "凭证号": (
                                    row.get("voucher_no", "")
                                    if source_name == "日记账"
                                    else ""
                                ),
                                "类型": self._match_type_name(
                                    candidate.match_type
                                ),
                                "处理状态": candidate.processing_status.value,
                                "纳入整池复核": (
                                    "是"
                                    if candidate.evidence.get(
                                        "included_in_pool_review",
                                        False,
                                    )
                                    else "否"
                                ),
                            }
                        )
        else:
            for source_name, frame in (
                ("银行流水", self.matcher.bank),
                ("日记账", self.matcher.journal),
            ):
                matched = frame[frame["matched"]].sort_values(
                    ["date", "amount_decimal"],
                )
                for index, row in matched.iterrows():
                    rows.append(
                        {
                            "匹配ID": row.get("match_id", ""),
                            "候选ID": "",
                            "来源": source_name,
                            "原文件行号": int(
                                row.get(
                                    "original_file_row",
                                    row.get("original_idx", index),
                                )
                            ),
                            "日期": row.get("date", ""),
                            "金额": float(row.get("amount", 0)),
                            "摘要": row.get("summary", ""),
                            "辅助文字": self._auxiliary_text(row),
                            "凭证号": (
                                row.get("voucher_no", "")
                                if source_name == "日记账"
                                else ""
                            ),
                            "类型": self._match_type_name(
                                str(row.get("match_type", ""))
                            ),
                            "处理状态": "自动确认",
                            "纳入整池复核": "否",
                        }
                    )
        return pd.DataFrame(rows, columns=columns)

    def _build_trivial_table(self) -> pd.DataFrame:
        columns = [
            "月份", "差异池", "池累计金额", "池是否超限", "匹配候选ID",
            "差异金额", "纳入整池复核", "处理状态", "处理依据",
        ]
        rows = []
        for pool in getattr(self.matcher, "difference_pools", []):
            for component in pool.components:
                rows.append(
                    {
                        "月份": pool.month,
                        "差异池": pool.pool_type.value,
                        "池累计金额": float(
                            PrecisionEngine.from_integer_li(
                                pool.total_diff_li
                            )
                        ),
                        "池是否超限": (
                            "是"
                            if pool.exceeds_performance_materiality
                            else "否"
                        ),
                        "匹配候选ID": component.candidate_id,
                        "差异金额": float(
                            PrecisionEngine.from_integer_li(
                                component.diff_li
                            )
                        ),
                        "纳入整池复核": (
                            "是"
                            if component.included_in_pool_review
                            else "否"
                        ),
                        "处理状态": pool.processing_status.value,
                        "处理依据": pool.processing_reason,
                    }
                )
        return pd.DataFrame(rows, columns=columns)

    def _source_snapshot(
        self,
        frame: pd.DataFrame,
        indexes: List[int],
        prefix: str,
    ) -> Dict[str, Any]:
        """把一侧多笔组成压缩成可直接在复核表查看的证据。"""
        valid_indexes = sorted(
            {
                int(index)
                for index in indexes
                if int(index) in frame.index
            }
        )
        if not valid_indexes:
            return {
                f"{prefix}原文件行号": "",
                f"{prefix}日期": "",
                f"{prefix}金额": 0.0,
                f"{prefix}逐笔金额": "",
                f"{prefix}摘要": "",
                f"{prefix}辅助文字": "",
            }
        selected = frame.loc[valid_indexes]
        return {
            f"{prefix}原文件行号": "；".join(
                str(
                    int(
                        row.get(
                            "original_file_row",
                            row.get("original_idx", index),
                        )
                    )
                )
                for index, row in selected.iterrows()
            ),
            f"{prefix}日期": "；".join(
                pd.Timestamp(value).strftime("%Y-%m-%d")
                for value in selected["date"]
            ),
            f"{prefix}金额": float(selected["amount"].sum()),
            f"{prefix}逐笔金额": "；".join(
                str(float(value)) for value in selected["amount"]
            ),
            f"{prefix}摘要": "；".join(
                str(value)
                for value in selected["summary"]
                if str(value).strip()
            ),
            f"{prefix}辅助文字": "；".join(
                self._auxiliary_text(row)
                for _, row in selected.iterrows()
                if self._auxiliary_text(row)
            ),
        }

    def _candidate_snapshots(
        self,
        candidate,
    ) -> Dict[str, Any]:
        snapshots = self._source_snapshot(
            self.matcher.bank,
            list(candidate.bank_idxs),
            "银行",
        )
        snapshots.update(
            self._source_snapshot(
                self.matcher.journal,
                list(candidate.journal_idxs),
                "日记账",
            )
        )
        return snapshots

    def _build_pending_review_table(self) -> pd.DataFrame:
        columns = [
            "复核事项ID", "事项类型", "月份", "匹配类型", "银行笔数",
            "日记账笔数", "组金额", "差异金额", "综合可信度", "原因",
            "金额分", "日期分", "文字分", "结构分", "重要性规则",
            "银行原文件行号", "日记账原文件行号", "银行日期",
            "日记账日期", "银行金额", "日记账金额", "银行逐笔金额",
            "日记账逐笔金额", "银行摘要", "日记账摘要", "银行辅助文字",
            "日记账辅助文字", "大模型判断", "差异池ID",
            "差异池累计金额", "匹配候选ID", "组成数量",
            "复核结论", "复核说明",
        ]
        rows = []
        config = self.matcher.config
        for candidate in getattr(
            self.matcher,
            "selected_candidates",
            [],
        ):
            if candidate.processing_status.value != "待人工复核":
                continue
            if candidate.evidence.get("included_in_pool_review", False):
                continue
            importance_rule = (
                "超过实际执行重要性水平"
                if candidate.metrics.group_amount_li
                > PrecisionEngine.to_integer_li(
                    config.performance_materiality
                )
                else (
                    "跨月多对多"
                    if candidate.is_cross_month_many_to_many
                    else "综合可信度未达到自动确认门槛"
                )
            )
            row = {
                    "复核事项ID": candidate.final_match_id,
                    "事项类型": "匹配组",
                    "月份": (
                        min(
                            candidate.bank_dates
                            + candidate.journal_dates
                        ).strftime("%Y-%m")
                        if candidate.bank_dates or candidate.journal_dates
                        else ""
                    ),
                    "匹配类型": self._match_type_name(
                        candidate.match_type
                    ),
                    "银行笔数": len(candidate.bank_idxs),
                    "日记账笔数": len(candidate.journal_idxs),
                    "组金额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.group_amount_li
                        )
                    ),
                    "差异金额": float(
                        PrecisionEngine.from_integer_li(
                            candidate.metrics.total_diff_li
                        )
                    ),
                    "综合可信度": candidate.scores.total,
                    "原因": candidate.processing_reason,
                    "金额分": candidate.scores.amount,
                    "日期分": candidate.scores.date,
                    "文字分": candidate.scores.text,
                    "结构分": candidate.scores.structure,
                    "重要性规则": importance_rule,
                    "大模型判断": (
                        candidate.llm_decision.reason
                        if candidate.llm_decision
                        else ""
                    ),
                    "差异池ID": "；".join(
                        str(value)
                        for value in candidate.evidence.get(
                            "difference_pool_ids",
                            [],
                        )
                    ),
                    "差异池累计金额": "",
                    "匹配候选ID": candidate.candidate_id,
                    "组成数量": (
                        len(candidate.bank_idxs)
                        + len(candidate.journal_idxs)
                    ),
                    "复核结论": "暂不处理",
                    "复核说明": "",
            }
            row.update(self._candidate_snapshots(candidate))
            rows.append(row)
        candidate_by_id = {
            candidate.candidate_id: candidate
            for candidate in getattr(
                self.matcher,
                "selected_candidates",
                [],
            )
        }
        for pool in getattr(self.matcher, "difference_pools", []):
            if not pool.exceeds_performance_materiality:
                continue
            component_candidates = [
                candidate_by_id[component.candidate_id]
                for component in pool.components
                if component.candidate_id in candidate_by_id
            ]
            bank_indexes = [
                index
                for candidate in component_candidates
                for index in candidate.bank_idxs
            ]
            journal_indexes = [
                index
                for candidate in component_candidates
                for index in candidate.journal_idxs
            ]
            row = {
                    "复核事项ID": pool.pool_id,
                    "事项类型": "月度差异池",
                    "月份": pool.month,
                    "匹配类型": pool.pool_type.value,
                    "银行笔数": "",
                    "日记账笔数": "",
                    "组金额": float(
                        PrecisionEngine.from_integer_li(
                            pool.total_diff_li
                        )
                    ),
                    "差异金额": float(
                        PrecisionEngine.from_integer_li(
                            pool.total_diff_li
                        )
                    ),
                    "综合可信度": "",
                    "原因": pool.processing_reason,
                    "金额分": "",
                    "日期分": "",
                    "文字分": "",
                    "结构分": "",
                    "重要性规则": "月度累计超过实际执行重要性水平",
                    "大模型判断": "；".join(
                        candidate.llm_decision.reason
                        for candidate in component_candidates
                        if candidate.llm_decision
                    ),
                    "差异池ID": pool.pool_id,
                    "差异池累计金额": float(
                        PrecisionEngine.from_integer_li(
                            pool.total_diff_li
                        )
                    ),
                    "匹配候选ID": "；".join(
                        component.candidate_id
                        for component in pool.components
                    ),
                    "组成数量": len(pool.components),
                    "复核结论": "暂不处理",
                    "复核说明": "",
            }
            row.update(
                self._source_snapshot(
                    self.matcher.bank,
                    bank_indexes,
                    "银行",
                )
            )
            row.update(
                self._source_snapshot(
                    self.matcher.journal,
                    journal_indexes,
                    "日记账",
                )
            )
            rows.append(row)
        rows.sort(
            key=lambda row: (
                str(row["月份"]),
                str(row["事项类型"]),
                str(row["复核事项ID"]),
            )
        )
        return pd.DataFrame(rows, columns=columns)

    def _build_unmatched_table(self, source: str) -> pd.DataFrame:
        frame = (
            self.matcher.bank
            if source == "bank"
            else self.matcher.journal
        )
        raw = self.raw_bank if source == "bank" else self.raw_journal
        unmatched = frame[~frame["matched"]]
        if raw is not None and not unmatched.empty:
            positions = [
                int(index) - 1
                for index in unmatched["original_idx"]
                if 0 <= int(index) - 1 < len(raw)
            ]
            result = raw.iloc[positions].copy()
            for column in result.columns:
                text = str(column).lower()
                if any(
                    keyword in text
                    for keyword in (
                        "金额", "amount", "发生额", "余额", "balance"
                    )
                ):
                    result[column] = result[column].map(
                        _restore_numeric_cells
                    )
            return result
        columns = ["日期", "金额", "摘要", "原文件行号"]
        if unmatched.empty:
            return pd.DataFrame(columns=columns)
        result = pd.DataFrame(
            {
                "日期": unmatched["date"],
                "金额": unmatched["amount"].map(float),
                "摘要": unmatched["summary"],
                "原文件行号": unmatched.get(
                    "original_file_row",
                    unmatched.get("original_idx", unmatched.index),
                ),
            }
        )
        return result.sort_values(
            ["日期", "金额"],
            kind="stable",
        )

    def _build_parameter_table(
        self,
        config: MatcherConfig,
        date_format: str,
    ) -> pd.DataFrame:
        rows = [
            ("实际执行重要性水平", float(config.performance_materiality)),
            (
                "明显微小错报临界值",
                float(config.clearly_trivial_threshold),
            ),
            ("自动确认最低综合可信度", config.auto_confirm_score),
            ("日期容差天数", config.tolerance_days),
            ("组合窗口天数", config.dfs_date_window),
            ("组合最大深度", config.max_dfs_depth),
            ("批量最少笔数", config.batch_min_count),
            ("最大候选数", config.max_candidates),
            ("是否允许异号", "是" if config.allow_mixed_sign else "否"),
            ("日期格式", date_format),
            (
                "请求随机种子",
                getattr(self.matcher, "run_parameters", {}).get(
                    "requested_random_seed",
                    config.random_seed,
                ),
            ),
            (
                "实际随机种子",
                getattr(self.matcher, "run_parameters", {}).get(
                    "actual_random_seed",
                    config.random_seed,
                ),
            ),
            ("运行时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        assistant = getattr(self.matcher, "llm_assistant", None)
        assistant_config = getattr(assistant, "config", None)
        rows.append(
            (
                "大模型辅助",
                "启用"
                if assistant_config is not None
                and getattr(assistant_config, "enabled", False)
                else "关闭",
            )
        )
        if assistant_config is not None and getattr(
            assistant_config,
            "enabled",
            False,
        ):
            rows.extend(
                [
                    ("大模型模式", getattr(assistant_config, "mode", "")),
                    ("大模型协议", getattr(assistant_config, "protocol", "")),
                    ("大模型模型", getattr(assistant_config, "model", "")),
                    (
                        "大模型服务地址",
                        sanitize_url(
                            getattr(assistant_config, "base_url", "")
                        ),
                    ),
                ]
            )
        return pd.DataFrame(rows, columns=["参数名称", "参数值"])

    def _build_llm_table(self) -> pd.DataFrame:
        columns = [
            "请求ID", "候选ID", "是否模型选择", "银行原文件行号",
            "日记账原文件行号", "银行日期", "日记账日期", "银行金额",
            "日记账金额", "银行摘要", "日记账摘要", "银行辅助文字",
            "日记账辅助文字", "实际发送字段", "服务", "调用协议",
            "接口地址", "模型", "本地综合可信度", "本地文字分",
            "金额分", "日期分", "文字分", "结构分", "模型语义分",
            "最终综合可信度", "判断理由", "支持证据", "冲突证据",
            "不确定性", "建议状态", "最终状态", "组金额", "差异金额",
            "实际执行重要性水平", "明显微小错报临界值",
            "自动确认最低综合可信度", "开始时间", "耗时毫秒", "用量",
            "是否降级", "错误", "脱敏原始回答",
        ]
        candidate_by_id = {
            candidate.candidate_id: candidate
            for candidate in getattr(self.matcher, "candidates", [])
        }
        selected_ids = {
            candidate.candidate_id
            for candidate in getattr(
                self.matcher,
                "selected_candidates",
                [],
            )
        }
        assistant_config = getattr(
            getattr(self.matcher, "llm_assistant", None),
            "config",
            None,
        )
        endpoint = sanitize_url(
            getattr(assistant_config, "base_url", "")
        )
        config = self.matcher.config
        rows = []
        for record in getattr(self.matcher, "llm_records", []):
            candidate_ids = tuple(record.candidate_ids)
            if not candidate_ids and record.selected_candidate_id:
                candidate_ids = (record.selected_candidate_id,)
            if not candidate_ids:
                candidate_ids = ("",)
            for candidate_id in candidate_ids:
                candidate = candidate_by_id.get(candidate_id)
                row = {
                    "请求ID": record.request_id,
                    "候选ID": candidate_id,
                    "是否模型选择": (
                        "是"
                        if candidate_id
                        and candidate_id
                        == record.selected_candidate_id
                        else "否"
                    ),
                    "实际发送字段": "；".join(record.sent_fields),
                    "服务": record.provider,
                    "调用协议": record.protocol,
                    "接口地址": endpoint,
                    "模型": record.model,
                    "本地综合可信度": (
                        candidate.evidence.get(
                            "pre_llm_total_score",
                            candidate.scores.total,
                        )
                        if candidate is not None
                        else ""
                    ),
                    "本地文字分": (
                        candidate.text_evidence.local_score
                        if candidate is not None
                        and candidate.text_evidence is not None
                        else ""
                    ),
                    "金额分": (
                        candidate.scores.amount
                        if candidate is not None
                        else ""
                    ),
                    "日期分": (
                        candidate.scores.date
                        if candidate is not None
                        else ""
                    ),
                    "文字分": (
                        candidate.scores.text
                        if candidate is not None
                        else ""
                    ),
                    "结构分": (
                        candidate.scores.structure
                        if candidate is not None
                        else ""
                    ),
                    "模型语义分": record.semantic_score,
                    "最终综合可信度": (
                        candidate.scores.total
                        if candidate is not None
                        else ""
                    ),
                    "判断理由": redact_sensitive_text(record.reason),
                    "支持证据": redact_sensitive_text(
                        "；".join(record.supporting_evidence)
                    ),
                    "冲突证据": redact_sensitive_text(
                        "；".join(record.conflicting_evidence)
                    ),
                    "不确定性": record.uncertainty,
                    "建议状态": record.suggested_status,
                    "最终状态": (
                        candidate.processing_status.value
                        if candidate is not None
                        and candidate_id in selected_ids
                        else ""
                    ),
                    "组金额": (
                        float(
                            PrecisionEngine.from_integer_li(
                                candidate.metrics.group_amount_li
                            )
                        )
                        if candidate is not None
                        else ""
                    ),
                    "差异金额": (
                        float(
                            PrecisionEngine.from_integer_li(
                                candidate.metrics.total_diff_li
                            )
                        )
                        if candidate is not None
                        else ""
                    ),
                    "实际执行重要性水平": float(
                        config.performance_materiality
                    ),
                    "明显微小错报临界值": float(
                        config.clearly_trivial_threshold
                    ),
                    "自动确认最低综合可信度": (
                        config.auto_confirm_score
                    ),
                    "开始时间": record.started_at,
                    "耗时毫秒": record.duration_ms,
                    "用量": json.dumps(
                        record.usage,
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                    "是否降级": "是" if record.fallback_used else "否",
                    "错误": redact_sensitive_text(record.error),
                    "脱敏原始回答": redact_sensitive_text(
                        record.raw_response
                    ),
                }
                if candidate is not None:
                    row.update(self._candidate_snapshots(candidate))
                rows.append(row)
        return pd.DataFrame(rows, columns=columns)

    def _build_balance_tables(
        self,
        *,
        bank_has_balance: bool,
        journal_has_balance: bool,
    ) -> dict[str, pd.DataFrame]:
        tables: dict[str, pd.DataFrame] = {}
        bank_recalculator = BalanceRecalculator()
        journal_recalculator = BalanceRecalculator()
        bank_balances = bank_recalculator.recalculate(self.matcher.bank)
        journal_balances = journal_recalculator.recalculate(
            self.matcher.journal
        )
        reconciler = BalanceReconciler(
            bank_balances=bank_balances,
            journal_balances=journal_balances,
        )
        differences = reconciler.generate_diff_report()
        if differences:
            tables["余额差异明细"] = pd.DataFrame(
                [
                    {
                        "日期": item.date,
                        "银行余额": float(item.bank_balance or 0),
                        "日记账余额": float(item.journal_balance or 0),
                        "差额": float(item.diff or 0),
                        "差异类型": item.diff_type,
                    }
                    for item in differences
                ]
            )
        continuity_rows = []
        if bank_has_balance:
            continuity_rows.extend(
                self.check_balance_continuity(
                    self.matcher.bank,
                    source="银行流水",
                )
            )
        if journal_has_balance:
            continuity_rows.extend(
                self.check_balance_continuity(
                    self.matcher.journal,
                    source="日记账",
                )
            )
        if continuity_rows:
            tables["余额连续性异常"] = pd.DataFrame(continuity_rows)
        return tables

    def build_report_tables(
        self,
        config: MatcherConfig,
        date_format: str = "auto",
    ) -> dict[str, pd.DataFrame]:
        """在写入 Excel 前构造所有可单独检查的结构化表。"""
        balance_possible, has_warning = self._prepare_initial_balance()
        bank_has_balance = self._has_balance_data(
            self.matcher.bank,
            self.bank_mapping,
        )
        journal_source = (
            self.raw_journal
            if self.raw_journal is not None
            else self.matcher.journal
        )
        journal_has_balance = self._has_balance_data(
            journal_source,
            self.journal_mapping,
        )
        daily, monthly = self._build_daily_and_monthly_tables()
        tables = {
            "核对汇总": self._build_summary_table(
                config,
                balance_check_possible=balance_possible,
                has_warning=has_warning,
            ),
            "每日统计": daily,
            "月度统计": monthly,
            "匹配明细": self._build_match_group_table(),
            "匹配组成明细": self._build_match_component_table(),
            "明显微小错报": self._build_trivial_table(),
            "待人工复核": self._build_pending_review_table(),
            "银行未达": self._build_unmatched_table("bank"),
            "日记账未达": self._build_unmatched_table("journal"),
            "运行参数": self._build_parameter_table(
                config,
                date_format,
            ),
        }
        if self.precheck_report is not None:
            tables["输入预检查"] = self.precheck_report.to_dataframe()
        tables.update(
            self._build_balance_tables(
                bank_has_balance=bank_has_balance,
                journal_has_balance=journal_has_balance,
            )
        )
        llm_table = self._build_llm_table()
        if not llm_table.empty:
            tables["大模型辅助明细"] = llm_table
        if self.error_collector and self.error_collector.has_errors():
            errors = self.error_collector.get_all_errors()
            if errors:
                tables["解析异常明细"] = pd.DataFrame(errors)
        return {
            name: self._safe_table(frame)
            for name, frame in tables.items()
        }

    def generate_report(
        self,
        output_path: str,
        config: Optional[MatcherConfig] = None,
        bank_path: Optional[str] = None,
        journal_path: Optional[str] = None,
        date_format: str = "auto",
    ) -> None:
        """生成组级清晰、可复核且防公式注入的 Excel 报告。"""
        del bank_path, journal_path
        effective_config = config or self.matcher.config
        tables = self.build_report_tables(
            effective_config,
            date_format=date_format,
        )
        make_excel(
            list(tables.items()),
            output_path,
            theme="deep-navy",
        )
        workbook = load_workbook(output_path)
        has_warning = bool(
            self.initial_balance_warning
            and self.initial_balance_warning.has_warning
        )
        if "核对汇总" in workbook.sheetnames:
            self._postprocess_summary(
                workbook["核对汇总"],
                has_warning,
            )
        if "匹配明细" in workbook.sheetnames:
            self._postprocess_details(workbook["匹配明细"])
        for sheet_name in (
            "每日统计",
            "月度统计",
            "余额差异明细",
            "余额连续性异常",
        ):
            if sheet_name in workbook.sheetnames:
                self._postprocess_diff_columns(workbook[sheet_name])

        if "待人工复核" in workbook.sheetnames:
            sheet = workbook["待人工复核"]
            header_columns = {
                cell.value: cell.column
                for cell in sheet[1]
                if cell.value is not None
            }
            conclusion_column = header_columns.get("复核结论")
            if conclusion_column is not None:
                validation = DataValidation(
                    type="list",
                    formula1='"接受,拒绝,暂不处理"',
                    allow_blank=False,
                )
                validation.error = "请选择接受、拒绝或暂不处理"
                validation.errorTitle = "复核结论无效"
                sheet.add_data_validation(validation)
                column_letter = get_column_letter(conclusion_column)
                validation.add(
                    f"{column_letter}2:"
                    f"{column_letter}{max(2, sheet.max_row)}"
                )
        workbook.save(output_path)
