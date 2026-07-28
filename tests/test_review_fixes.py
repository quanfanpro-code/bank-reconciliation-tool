"""
2026-07-28 全面复核修复 — 回归测试
每个用例对应一个已实证的缺陷，先红后绿。
"""
from decimal import Decimal

import pandas as pd
import pytest
from openpyxl import load_workbook

from balance import BalanceRecalculator
from data_loader import DataLoader, ParseErrorCollector
from data_structures import MatcherConfig
from matcher import Matcher
from reporter import Reporter
from validate import validate_config_params


# ==========================================
# Task 1: validate 上限与随机种子
# ==========================================

def _valid_kwargs(**overrides):
    kwargs = dict(
        tolerance_days="31", dfs_window="31", dfs_depth="30", greedy_attempts="3",
        random_seed="0", similarity_threshold="0.5", similarity_high="0.7",
        max_candidates="30", memory_limit="6.0", bank_skip="0", journal_skip="0",
    )
    kwargs.update(overrides)
    return kwargs


def test_GUI默认参数能通过校验():
    """出厂默认参数（窗口31/深度30）不得被自己的校验拒绝"""
    ok, msg = validate_config_params(**_valid_kwargs())
    assert ok, msg


def test_随机种子负一合法():
    """readme 文档化语义：-1 = 每次运行使用不同随机种子"""
    ok, msg = validate_config_params(**_valid_kwargs(random_seed="-1"))
    assert ok, msg


def test_组合窗口上限仍然生效():
    ok, _ = validate_config_params(**_valid_kwargs(dfs_window="91"))
    assert not ok


def test_最大深度上限仍然生效():
    ok, _ = validate_config_params(**_valid_kwargs(dfs_depth="101"))
    assert not ok


# ==========================================
# Task 3: 批量聚合 category 摘要崩溃
# ==========================================

def test_批量聚合_category摘要不崩溃且匹配():
    """summary 被 _downcast_dtypes 转 category 后，fillna('') 会抛 TypeError"""
    n = 12
    bank = pd.DataFrame({
        "date": [pd.Timestamp("2024-01-05")] * n,
        # 真实数据中常有个别空摘要；category + NaN + fillna('') 会抛 TypeError
        "summary": pd.Categorical(["代发工资"] * (n - 1) + [None]),
        "amount_decimal": [500000] * n,  # 50.00 元 = 500000 厘
    })
    journal = pd.DataFrame({
        "date": [pd.Timestamp("2024-01-06")],
        "summary": ["代发工资总额"],
        "amount_decimal": [500000 * n],
    })

    matcher = Matcher(bank, journal, MatcherConfig())
    matcher.match_batch_aggregation()

    assert matcher.bank["matched"].sum() == n
    assert matcher.journal["matched"].sum() == 1


# ==========================================
# Task 4: 空余额语义
# ==========================================

def test_空余额单元格为None而非零():
    loader = DataLoader()
    df = pd.DataFrame({
        "日期": ["2024-01-01", "2024-01-02"],
        "金额": [100, -50], "摘要": ["a", "b"], "余额": [1000.0, None],
    })
    mapping = {"date": "日期", "amount": "金额", "summary": "摘要",
               "voucher": None, "balance": "余额", "mode": "signed_amount"}

    res = loader.standardize_data(df, mapping, "bank", "auto")
    by_row = res.set_index("original_idx")["balance"]

    assert by_row.loc[1] == Decimal("1000.00")
    assert pd.isna(by_row.loc[2])


# ==========================================
# Task 5: 期初余额提取
# ==========================================

def test_期初提取_千分位余额不崩溃():
    """原始数据余额列为千分位字符串时不得抛 InvalidOperation"""
    df = pd.DataFrame({
        "交易日期": ["2024-01-01", "2024-01-02"],
        "业务说明": ["收款", "付款"],
        "账户余额": ["1,234.56", "1,134.56"],
    })
    mapping = {"date": "交易日期", "summary": "业务说明", "balance": "账户余额",
               "debit": None, "credit": None, "amount": None, "mode": "debit_credit"}

    result = BalanceRecalculator.extract_initial_balance(df, mapping)

    assert isinstance(result, Decimal)


def test_期初提取_mapping列名优先于猜列():
    """mapping 指定的列名不在内置猜测词表中时也必须生效"""
    df = pd.DataFrame({
        "交易日期": ["2024-01-01", "2024-01-01"],
        "业务说明": ["收款", "付款"],
        "账户余额": ["1000.00", "900.00"],
        "收入": ["200.00", ""],
        "支出": ["", "100.00"],
    })
    mapping = {"date": "交易日期", "summary": "业务说明", "balance": "账户余额",
               "debit": "收入", "credit": "支出", "amount": None, "mode": "debit_credit"}

    # 期初 = 首笔余额 - 首笔净额 = 1000 - (200 - 0) = 800
    result = BalanceRecalculator.extract_initial_balance(df, mapping)

    assert result == Decimal("800.00")


def test_期初推断_跳过空余额单元格():
    """标准化数据中首笔 balance 为 None 时，应使用首个有效余额行回补推算"""
    df = pd.DataFrame({
        "date": [pd.Timestamp("2024-01-01"), pd.Timestamp("2024-01-02")],
        "amount": [Decimal("100.00"), Decimal("-50.00")],
        "balance": [None, Decimal("950.00")],
        "original_idx": [1, 2],
    })

    # 首个有效余额 950（1-02 末） - 截至该日累计净额 (100-50)=50 → 期初 900
    result = BalanceRecalculator.extract_initial_balance(df)

    assert result == Decimal("900.00")


# ==========================================
# Task 6-8: Reporter 集成（generate_report 真实产出）
# ==========================================

def _std_df(rows):
    """构造与 standardize_data 输出同构的最小标准化 DataFrame"""
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    if "amount_decimal" not in df.columns:
        df["amount_decimal"] = df["amount"].apply(
            lambda a: int((a * 10000).quantize(Decimal("1"))))
    if "original_idx" not in df.columns:
        df["original_idx"] = range(1, len(df) + 1)
    if "original_file_row" not in df.columns:
        df["original_file_row"] = df["original_idx"] + 1
    if "voucher_no" not in df.columns:
        df["voucher_no"] = ""
    return df


def _read_sheet_rows(path, sheet_name):
    wb = load_workbook(path)
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v is not None for v in row):
            rows.append(row)
    return rows


def test_期初警告_无余额数据时跳过(tmp_path):
    """任一侧无有效余额列数据时：不发期初警告，汇总表注明跳过"""
    bank = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收款"], "balance": [None],
    })
    journal = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收货款"], "balance": [Decimal("500.00")],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    assert reporter.initial_balance_warning.has_warning is False
    texts = [str(v) for row in _read_sheet_rows(out, "核对汇总") for v in row]
    assert any("跳过期初核对" in t for t in texts)


def test_余额连续性异常sheet会生成(tmp_path):
    """readme 承诺的第六个 sheet：余额跳变必须被检出并写入独立 sheet"""
    bank = _std_df({
        "date": ["2024-01-01", "2024-01-02"],
        "amount": [Decimal("100.00"), Decimal("50.00")],
        "summary": ["收款", "收款"],
        "balance": [Decimal("1000.00"), Decimal("1200.00")],  # 预期 1050，跳变 150
    })
    journal = _std_df({
        "date": ["2024-01-01", "2024-01-02"],
        "amount": [Decimal("100.00"), Decimal("50.00")],
        "summary": ["收货款", "收货款"],
        "balance": [Decimal("500.00"), Decimal("550.00")],  # 连续，无异常
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    rows = _read_sheet_rows(out, "余额连续性异常")
    header = rows[0]
    src_col = header.index("来源")
    sources = [r[src_col] for r in rows[1:]]
    assert "银行流水" in sources
    assert "日记账" not in sources


def test_汇总表数值项为数字类型(tmp_path):
    """核对汇总的数值项必须是 int/float，不得在 Excel 中变成文本"""
    bank = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收款"], "balance": [None],
    })
    journal = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收货款"], "balance": [None],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    wb = load_workbook(out)
    ws = wb["核对汇总"]
    label_value = {}
    for row in ws.iter_rows(min_col=2, max_col=3, values_only=True):
        label_value[row[0]] = row[1]
    assert isinstance(label_value["银行流水总笔数"], (int, float))
    assert isinstance(label_value["银行流水总金额"], (int, float))


def test_匹配明细按日期金额排序可复现(tmp_path):
    """匹配明细行序必须确定性排序，不依赖 set 迭代顺序"""
    n = 8
    bank = _std_df({
        "date": [f"2024-01-{d:02d}" for d in range(1, n + 1)],
        "amount": [Decimal("100.00")] * n,
        "summary": ["收款"] * n,
        "balance": [None] * n,
    })
    journal = _std_df({
        "date": [f"2024-01-{d:02d}" for d in range(1, n + 1)],
        "amount": [Decimal("100.00")] * n,
        "summary": ["收货款"] * n,
        "balance": [None] * n,
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    # 故意乱序标记匹配（01-08 最先，01-01 最后）
    for i in reversed(range(n)):
        matcher._mark_matched([bank.index[i]], [journal.index[i]], "exact_1to1", "高")

    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"
    reporter.generate_report(str(out), config=MatcherConfig())

    rows = _read_sheet_rows(out, "匹配明细")
    header = rows[0]
    date_col = header.index("银_日期")
    dates = [r[date_col] for r in rows[1:]]
    assert dates == sorted(dates)
    assert dates[0].day == 1 and dates[-1].day == n


def test_期初警告涂色不覆盖表头(tmp_path):
    """期初警告高亮范围应为数据行 2..7，表头（第1行）不得被涂黄"""
    bank = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收款"], "balance": [Decimal("10000.00")],
    })
    journal = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收货款"], "balance": [Decimal("500.00")],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())
    assert reporter.initial_balance_warning.has_warning is True

    wb = load_workbook(out)
    ws = wb["核对汇总"]
    header_fill = ws.cell(row=1, column=2).fill
    assert header_fill.start_color.rgb != "00FFFF00" and header_fill.start_color.rgb != "FFFFFF00"
    assert ws.cell(row=2, column=2).fill.start_color.rgb in ("00FFFF00", "FFFFFF00")


# ==========================================
# Task 9: P2 组
# ==========================================

_NOW = pd.Timestamp("2024-01-15")


def test_随机贪心_洗牌真正生效():
    """修复前 shuffle 后被 sort 中和：排序贪心找不到的组合永远找不到"""
    from matcher import _randomized_greedy
    # 排序贪心：4(取) -> 3(跳过) -> 2(跳过) = 4 ≠ 5，必然失败；
    # 洗牌后 3+2=5 可被找到
    result = _randomized_greedy(
        window_amounts=[4, 3, 2],
        window_dates=[_NOW] * 3,
        window_indices=[0, 1, 2],
        target=5,
        num_attempts=20,
        random_seed=42,
    )
    assert result is not None
    idxs, _confidence = result
    assert sum([4, 3, 2][i] for i in idxs) == 5


def test_随机贪心_同种子结果可复现():
    from matcher import _randomized_greedy
    kwargs = dict(
        window_amounts=[4, 3, 2, 6, 1],
        window_dates=[_NOW] * 5,
        window_indices=[0, 1, 2, 3, 4],
        target=7, num_attempts=10, random_seed=7,
    )
    assert _randomized_greedy(**kwargs) == _randomized_greedy(**kwargs)


def test_方向列缺失时不得静默按正数():
    """single_amount_with_direction 模式缺少方向列必须显式报错"""
    loader = DataLoader()
    df = pd.DataFrame({
        "日期": ["2024-01-01"], "金额": ["100"], "摘要": ["a"],
    })
    mapping = {"date": "日期", "amount": "金额", "direction": None,
               "summary": "摘要", "voucher": None, "balance": None,
               "mode": "single_amount_with_direction"}

    with pytest.raises(ValueError, match="方向"):
        loader.standardize_data(df, mapping, "bank", "auto")


def test_round_decimal尊重decimals参数():
    from utils import round_decimal
    assert round_decimal(1.2345, 3) == 1.234 or round_decimal(1.2345, 3) == 1.235
    assert round_decimal(1.2345, 3) != round_decimal(1.2345, 2)
    assert round_decimal(2.675) == 2.68  # 默认 2 位行为不变（HALF_UP）


def test_方向列支持收付标注():
    """银行流水最常用的"收/付"方向标注必须被识别，支出行不得被静默丢弃"""
    from data_loader import ParseErrorCollector
    loader = DataLoader()
    ec = ParseErrorCollector()
    loader.error_collector = ec
    df = pd.DataFrame({
        "交易日期": ["2026-01-05", "2026-01-06"],
        "摘要": ["收入一笔", "支出一笔"],
        "金额": [100.0, 40.0],
        "方向": ["收", "付"],
    })
    mapping = {"date": "交易日期", "amount": "金额", "direction": "方向",
               "summary": "摘要", "voucher": None, "balance": None,
               "mode": "single_amount_with_direction"}
    out = loader.standardize_data(df, mapping, "bank", "auto")
    assert len(out) == 2, f"支出行被静默丢弃: 仅剩 {len(out)} 行"
    by_summary = dict(zip(out["summary"], out["amount"]))
    assert float(by_summary["收入一笔"]) == 100.0
    assert float(by_summary["支出一笔"]) == -40.0
    assert not ec.has_errors(), "合法方向值'付'被误记为解析异常"


def test_未达明细数值还原():
    """未达 sheet 的金额/余额列清洗后应还原为数值类型，Excel 中可直接求和"""
    from reporter import _restore_numeric_cells
    assert _restore_numeric_cells("500") == 500.0
    assert _restore_numeric_cells("107,350") == 107350.0
    assert _restore_numeric_cells("-1,234.56") == -1234.56
    assert _restore_numeric_cells("B公司货款") == "B公司货款"
    assert _restore_numeric_cells("") == ""
    assert _restore_numeric_cells(None) is None


def test_余额连续性_跨空余额日累计检查(tmp_path):
    """空余额日不应让检查断链：其后首个非空日按"最近有效余额+累计净额"校验"""
    bank = _std_df({
        "date": ["2026-01-05", "2026-01-10", "2026-01-15", "2026-01-20"],
        "amount": [Decimal("10000.00"), Decimal("-3000.00"), Decimal("50.00"), Decimal("-200.00")],
        "summary": ["货款", "工资", "利息", "服务费"],
        # 01-15 余额为空；01-20 应为 107000+50-200=106850，故意写成 106900（差 50 元）
        "balance": [Decimal("110000.00"), Decimal("107000.00"), None, Decimal("106900.00")],
    })
    journal = _std_df({
        "date": ["2026-01-05"],
        "amount": [Decimal("10000.00")],
        "summary": ["货款"],
        "balance": [Decimal("110000.00")],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    rows = _read_sheet_rows(out, "余额连续性异常")
    header = rows[0]
    src_col = header.index("来源")
    date_col = header.index("日期")
    hits = [r for r in rows[1:] if r[src_col] == "银行流水" and "2026-01-20" in str(r[date_col])]
    assert len(hits) == 1, f"跨空余额日的余额错误未被检出: {rows[1:]}"


def test_汇总表期初余额为数值类型(tmp_path):
    """期初余额在 Excel 中必须是数值（千分位显示由单元格格式负责，值可参与计算）"""
    bank = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收款"], "balance": [Decimal("1100.00")],
    })
    journal = _std_df({
        "date": ["2024-01-01"], "amount": [Decimal("100.00")],
        "summary": ["收货款"], "balance": [Decimal("1100.00")],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    found = 0
    for row in _read_sheet_rows(out, "核对汇总"):
        for i, v in enumerate(row[:-1]):
            if isinstance(v, str) and v in ("银行期初余额", "日记账期初余额", "期初余额差额"):
                cell_val = row[i + 1]
                assert isinstance(cell_val, (int, float)) and not isinstance(cell_val, bool), \
                    f"{v} 应为数值类型，实际: {type(cell_val).__name__}={cell_val!r}"
                found += 1
    assert found == 3, f"期初区块行数异常: {found}"


def test_连续性sheet区间净额为数值非日期(tmp_path):
    """"区间净额"列不得被列名推断误判为日期格式（负净额会变成 1899 年的日期）"""
    bank = _std_df({
        "date": ["2026-01-05", "2026-01-10", "2026-01-15", "2026-01-20"],
        "amount": [Decimal("10000.00"), Decimal("-3000.00"), Decimal("50.00"), Decimal("-200.00")],
        "summary": ["货款", "工资", "利息", "服务费"],
        "balance": [Decimal("110000.00"), Decimal("107000.00"), None, Decimal("106900.00")],
    })
    journal = _std_df({
        "date": ["2026-01-05"], "amount": [Decimal("10000.00")],
        "summary": ["货款"], "balance": [Decimal("110000.00")],
    })
    matcher = Matcher(bank, journal, MatcherConfig())
    reporter = Reporter(matcher)
    out = tmp_path / "r.xlsx"

    reporter.generate_report(str(out), config=MatcherConfig())

    rows = _read_sheet_rows(out, "余额连续性异常")
    header = rows[0]
    net_col = header.index("区间净额")
    val = rows[1][net_col]
    assert isinstance(val, (int, float)) and not isinstance(val, bool), \
        f"区间净额应为数值，实际: {type(val).__name__}={val!r}"
    assert float(val) == -150.0


def test_余额列推断为金额格式():
    """"余额"类列名应命中 money 关键词（readme 承诺所有金额千分位格式）"""
    from make_excel import _infer_column_type
    s = pd.Series([Decimal("107000.00")])
    assert _infer_column_type("银行余额", s) == 'money'
    assert _infer_column_type("基准余额", s) == 'money'
    assert _infer_column_type("期末余额", s) == 'money'
