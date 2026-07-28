from decimal import Decimal

import pandas as pd
from openpyxl import load_workbook

from data_structures import LLMDecisionRecord, MatcherConfig
from matcher import Matcher
from precision_engine import PrecisionEngine
from reporter import Reporter


def _report_df(date, amounts, source, summaries=None):
    summaries = summaries or [""] * len(amounts)
    records = [
            {
                "date": pd.Timestamp(date),
                "amount": Decimal(str(amount)),
                "amount_decimal": PrecisionEngine.to_integer_li(amount),
                "summary": summaries[index],
                "aux_text_fields": {"摘要": summaries[index]},
                "balance": None,
                "voucher_no": f"记-{index + 1:03d}" if source == "journal" else "",
                "source": source,
                "original_idx": index + 1,
                "original_file_row": index + 2,
            }
            for index, amount in enumerate(amounts)
        ]
    return pd.DataFrame(
        records,
        columns=[
            "date",
            "amount",
            "amount_decimal",
            "summary",
            "aux_text_fields",
            "balance",
            "voucher_no",
            "source",
            "original_idx",
            "original_file_row",
        ],
    )


def _three_vs_five_reporter():
    matcher = Matcher(
        _report_df(
            "2026-01-15",
            [200, 200, 100],
            "bank",
            ["销售回款"] * 3,
        ),
        _report_df(
            "2026-01-15",
            [100, 100, 100, 100, 100],
            "journal",
            ["销售回款"] * 5,
        ),
        MatcherConfig(),
    )
    matcher.run()
    return Reporter(matcher)


def _pending_reporter():
    matcher = Matcher(
        _report_df(
            "2026-01-15",
            [120000],
            "bank",
            ["设备款"],
        ),
        _report_df(
            "2026-01-15",
            [120000],
            "journal",
            ["设备款"],
        ),
        MatcherConfig(),
    )
    matcher.run()
    return Reporter(matcher)


def test_多对多报告按组展示且组成明细不伪造一一对应(tmp_path):
    output = tmp_path / "report.xlsx"
    _three_vs_five_reporter().generate_report(
        str(output),
        config=MatcherConfig(),
    )

    groups = pd.read_excel(output, sheet_name="匹配明细")
    components = pd.read_excel(output, sheet_name="匹配组成明细")

    assert len(groups) == 1
    assert len(components) == 8
    assert set(components["来源"]) == {"银行流水", "日记账"}
    assert pd.api.types.is_numeric_dtype(components["金额"])


def test_待人工复核提供三个固定下拉选项和说明列(tmp_path):
    output = tmp_path / "report.xlsx"
    _pending_reporter().generate_report(
        str(output),
        config=MatcherConfig(),
    )

    workbook = load_workbook(output)
    sheet = workbook["待人工复核"]
    validations = list(sheet.data_validations.dataValidation)
    headers = [cell.value for cell in sheet[1]]

    assert '"接受,拒绝,暂不处理"' in {
        item.formula1 for item in validations
    }
    assert "复核说明" in headers
    conclusion_column = headers.index("复核结论") + 1
    assert sheet.cell(row=2, column=conclusion_column).value == "暂不处理"
    assert sheet.auto_filter.ref is not None
    assert sheet.auto_filter.ref.startswith("B1:")


def test_待人工复核直接提供做判断所需的完整证据():
    pending = _pending_reporter().build_report_tables(
        MatcherConfig()
    )["待人工复核"]

    assert {
        "银行原文件行号",
        "日记账原文件行号",
        "银行日期",
        "日记账日期",
        "银行金额",
        "日记账金额",
        "银行辅助文字",
        "日记账辅助文字",
        "金额分",
        "日期分",
        "文字分",
        "结构分",
        "重要性规则",
        "大模型判断",
        "差异池ID",
        "差异池累计金额",
    }.issubset(pending.columns)
    row = pending.iloc[0]
    assert "2" in str(row["银行原文件行号"])
    assert "2" in str(row["日记账原文件行号"])
    assert float(row["银行金额"]) == 120000
    assert float(row["日记账金额"]) == 120000
    assert "设备款" in str(row["银行辅助文字"])
    assert row["重要性规则"] == "超过实际执行重要性水平"


def test_精确匹配率与自动处理率分开且数值不超过百分之百():
    tables = _three_vs_five_reporter().build_report_tables(
        MatcherConfig()
    )
    summary = dict(
        zip(tables["核对汇总"]["项目"], tables["核对汇总"]["数值"])
    )

    assert isinstance(summary["精确匹配率"], (int, float))
    assert isinstance(summary["自动处理率"], (int, float))
    assert 0 <= summary["精确匹配率"] <= 1
    assert 0 <= summary["自动处理率"] <= 1
    assert 0 <= summary["低可信度组占比"] <= 1
    assert {
        "自动确认组数",
        "自动确认金额",
        "明显微小错报组数",
        "明显微小错报金额",
        "待人工复核事项数",
        "待人工复核金额",
        "银行未找到候选笔数",
        "银行未找到候选金额",
        "日记账未找到候选笔数",
        "日记账未找到候选金额",
        "大模型成功次数",
        "大模型降级次数",
    }.issubset(summary)


def test_匹配组报告披露双方收支净额差异池和模型参与情况():
    groups = _three_vs_five_reporter().build_report_tables(
        MatcherConfig()
    )["匹配明细"]

    assert {
        "银行收入",
        "银行支出",
        "银行净额",
        "日记账收入",
        "日记账支出",
        "日记账净额",
        "差异池ID",
        "是否使用大模型",
    }.issubset(groups.columns)


def test_空数据仍生成所有固定工作表和运行参数(tmp_path):
    empty = _report_df("2026-01-01", [], "bank")
    matcher = Matcher(empty, empty, MatcherConfig())
    reporter = Reporter(matcher)
    output = tmp_path / "empty.xlsx"

    reporter.generate_report(str(output), config=MatcherConfig())
    workbook = load_workbook(output)

    required = {
        "核对汇总",
        "每日统计",
        "月度统计",
        "匹配明细",
        "匹配组成明细",
        "明显微小错报",
        "待人工复核",
        "银行未达",
        "日记账未达",
        "运行参数",
    }
    assert required.issubset(workbook.sheetnames)


def test_用户文本不能被Excel当作公式执行(tmp_path):
    dangerous = ["=1+1", "+CMD", "-2+3", "@SUM(A1)", "\t=2+2"]
    bank = _report_df(
        "2026-01-01",
        [1, 2, 3, 4, 5],
        "bank",
        dangerous,
    )
    journal = _report_df("2026-01-01", [], "journal")
    reporter = Reporter(Matcher(bank, journal, MatcherConfig()))
    output = tmp_path / "safe.xlsx"

    reporter.generate_report(str(output), config=MatcherConfig())
    workbook = load_workbook(output, data_only=False)

    sheet = workbook["银行未达"]
    summary_column = next(
        cell.column
        for cell in sheet[1]
        if cell.value == "摘要"
    )
    for row_number in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_number, column=summary_column)
        assert cell.data_type != "f"


def test_大模型明细字段完整且敏感内容被清理(tmp_path):
    matcher = _pending_reporter().matcher
    candidate = matcher.selected_candidates[0]
    matcher.llm_records = [
        LLMDecisionRecord(
            request_id="R1",
            candidate_ids=(candidate.candidate_id,),
            sent_fields=("日期", "金额", "摘要"),
            selected_candidate_id=candidate.candidate_id,
            semantic_score=88,
            reason="账号622233334444支持匹配",
            supporting_evidence=("摘要一致",),
            conflicting_evidence=(),
            uncertainty="低",
            suggested_status="自动确认",
            provider="在线 API",
            model="mock-model",
            started_at="2026-01-01T10:00:00+08:00",
            duration_ms=120,
            usage={"input_tokens": 10},
            raw_response=(
                "账号622233334444；"
                "https://example.com/v1?api_key=secret-key"
            ),
        )
    ]
    output = tmp_path / "llm.xlsx"

    Reporter(matcher).generate_report(
        str(output),
        config=MatcherConfig(),
    )
    table = pd.read_excel(output, sheet_name="大模型辅助明细")
    body = table.to_string()

    assert {
        "本地文字分",
        "候选ID",
        "是否模型选择",
        "银行原文件行号",
        "日记账原文件行号",
        "银行日期",
        "日记账日期",
        "银行金额",
        "日记账金额",
        "实际发送字段",
        "接口地址",
        "本地综合可信度",
        "模型语义分",
        "判断理由",
        "支持证据",
        "冲突证据",
        "不确定性",
        "最终状态",
        "服务",
        "模型",
        "耗时毫秒",
        "用量",
        "是否降级",
        "实际执行重要性水平",
        "明显微小错报临界值",
        "自动确认最低综合可信度",
        "脱敏原始回答",
    }.issubset(table.columns)
    assert str(candidate.bank_idxs[0] + 2) in str(
        table.iloc[0]["银行原文件行号"]
    )
    assert "622233334444" not in body
    assert "secret-key" not in body
