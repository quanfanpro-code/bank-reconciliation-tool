from pathlib import Path
from decimal import Decimal

import pandas as pd
import pytest
from openpyxl import Workbook

import input_precheck
from application import run_reconciliation
from data_loader import DataLoader
from data_structures import MatcherConfig
import gui


def _build_merged_header_workbook(tmp_path: Path) -> Path:
    path = tmp_path / "合并表头.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["银行交易明细"])
    sheet.merge_cells("A1:E1")
    sheet.append(["日期", "发生额", None, "辅助信息", None])
    sheet.append([None, "借方", "贷方", "摘要", "对方户名"])
    sheet.merge_cells("A2:A3")
    sheet.merge_cells("B2:C2")
    sheet.merge_cells("D2:E2")
    sheet.append(["2026-01-01", 0, 100, "销售收款", "甲公司"])
    sheet.append(["2026-01-02", 20, 0, "支付手续费", "银行"])
    workbook.save(path)
    return path


def test_普通一行表头保持兼容(tmp_path):
    path = tmp_path / "普通表头.xlsx"
    pd.DataFrame(
        [{"日期": "2026-01-01", "收入": 100, "支出": 0, "摘要": "收款"}]
    ).to_excel(path, index=False)

    loader = DataLoader()
    structure = loader.detect_table_structure(path)
    data = loader.load_file(
        path,
        skiprows=structure.skiprows,
        header_rows=structure.header_rows,
        derived_columns=structure.columns,
    )

    assert structure.skiprows == 0
    assert structure.header_rows == 1
    assert structure.columns == ["日期", "收入", "支出", "摘要"]
    assert data.iloc[0].to_dict() == {
        "日期": "2026-01-01",
        "收入": 100,
        "支出": 0,
        "摘要": "收款",
    }


def test_标题与两行合并表头生成稳定无重复列名(tmp_path):
    path = _build_merged_header_workbook(tmp_path)

    loader = DataLoader()
    structure = loader.detect_table_structure(path)
    data = loader.load_file(
        path,
        skiprows=structure.skiprows,
        header_rows=structure.header_rows,
        derived_columns=structure.columns,
    )

    assert structure.skiprows == 1
    assert structure.header_rows == 2
    assert structure.columns == [
        "日期",
        "发生额｜借方",
        "发生额｜贷方",
        "辅助信息｜摘要",
        "辅助信息｜对方户名",
    ]
    assert len(structure.columns) == len(set(structure.columns))
    assert data.shape == (2, 5)
    assert data.iloc[0]["辅助信息｜对方户名"] == "甲公司"


def test_多行表头后的原文件行号保持可追溯(tmp_path):
    path = _build_merged_header_workbook(tmp_path)
    loader = DataLoader()
    structure = loader.detect_table_structure(path)
    data = loader.load_file(
        path,
        skiprows=structure.skiprows,
        header_rows=structure.header_rows,
        derived_columns=structure.columns,
    )
    mapping = {
        "date": "日期",
        "debit": "发生额｜借方",
        "credit": "发生额｜贷方",
        "summary": "辅助信息｜摘要",
        "auxiliary_text_columns": ["辅助信息｜摘要", "辅助信息｜对方户名"],
        "mode": "debit_credit",
    }

    standardized = loader.standardize_data(
        data,
        mapping,
        "bank",
        skiprows_offset=structure.skiprows,
        header_rows=structure.header_rows,
    )

    assert standardized["original_file_row"].tolist() == [4, 5]


def test_重复表头派生列名按出现顺序去重(tmp_path):
    path = tmp_path / "重复列名.csv"
    path.write_text(
        "日期,金额,金额,摘要\n2026-01-01,10,20,测试\n",
        encoding="utf-8-sig",
    )

    structure = DataLoader().detect_table_structure(path)

    assert structure.columns == ["日期", "金额", "金额_2", "摘要"]
    assert structure.columns == DataLoader().detect_table_structure(path).columns


def _structure(columns, *, ambiguous=False, candidates=()):
    return input_precheck.TableStructure(
        skiprows=0,
        header_rows=1,
        columns=list(columns),
        score=10,
        ambiguous=ambiguous,
        candidates=tuple(candidates),
        explanation="测试结构",
    )


def _standardized(rows, source):
    return pd.DataFrame(
        [
            {
                "date": pd.Timestamp(date),
                "amount": Decimal(str(amount)),
                "summary": summary,
                "source": source,
            }
            for date, amount, summary in rows
        ]
    )


def _signed_mapping():
    return {
        "date": "日期",
        "amount": "金额",
        "summary": "摘要",
        "auxiliary_text_columns": ["摘要", "对方户名"],
        "mode": "signed_amount",
    }


def _build_report(
    raw_bank,
    raw_journal,
    bank,
    journal,
    *,
    bank_mapping=None,
    journal_mapping=None,
    bank_structure=None,
    journal_structure=None,
    parse_errors=(),
):
    mapping = _signed_mapping()
    return input_precheck.build_input_precheck(
        raw_bank=raw_bank,
        raw_journal=raw_journal,
        bank=bank,
        journal=journal,
        bank_mapping=bank_mapping or mapping,
        journal_mapping=journal_mapping or mapping,
        bank_structure=bank_structure or _structure(raw_bank.columns),
        journal_structure=journal_structure or _structure(raw_journal.columns),
        parse_errors=list(parse_errors),
    )


def test_日期范围与收支合计分别比较且只提示():
    raw_bank = pd.DataFrame(
        [
            {"日期": "2026-01-01", "金额": 100, "摘要": "收款", "对方户名": "甲"},
            {"日期": "2026-01-03", "金额": -30, "摘要": "付款", "对方户名": "乙"},
        ]
    )
    raw_journal = pd.DataFrame(
        [
            {"日期": "2026-01-02", "金额": 80, "摘要": "收款", "对方户名": "甲"},
            {"日期": "2026-01-04", "金额": -20, "摘要": "付款", "对方户名": "乙"},
        ]
    )

    report = _build_report(
        raw_bank,
        raw_journal,
        _standardized(
            [("2026-01-01", 100, "收款"), ("2026-01-03", -30, "付款")],
            "bank",
        ),
        _standardized(
            [("2026-01-02", 80, "收款"), ("2026-01-04", -20, "付款")],
            "journal",
        ),
    )

    date_item = next(item for item in report.items if item.name == "日期范围")
    amount_item = next(item for item in report.items if item.name == "金额合计")
    assert report.has_blockers is False
    assert date_item.status == "提示"
    assert amount_item.status == "提示"
    assert amount_item.bank_result == "收入 100.00；支出 30.00"
    assert amount_item.journal_result == "收入 80.00；支出 20.00"
    assert amount_item.comparison == "收入差额 20.00；支出差额 10.00"


def test_无效方向与必填列缺失会阻止运行():
    raw = pd.DataFrame(
        [{"日期": "2026-01-01", "金额": 100, "摘要": "测试", "对方户名": "甲"}]
    )
    mapping = {
        "date": "日期",
        "amount": "金额",
        "direction": "方向",
        "summary": "摘要",
        "mode": "single_amount_with_direction",
    }
    report = _build_report(
        raw,
        raw,
        _standardized([("2026-01-01", 100, "测试")], "bank"),
        _standardized([("2026-01-01", 100, "测试")], "journal"),
        bank_mapping=mapping,
        journal_mapping=mapping,
        parse_errors=[
            {
                "type": "方向解析失败",
                "source_type": "bank",
                "row": 2,
                "original_value": "未知",
                "column": "方向",
            }
        ],
    )

    assert report.has_blockers is True
    assert next(item for item in report.items if item.name == "金额方向").status == "阻止"
    assert next(item for item in report.items if item.name == "必填字段").status == "阻止"
    assert "方向" in report.blocker_message()


def test_少量解析异常非交易行和低文字非空率只提示():
    raw_bank = pd.DataFrame(
        [
            {"日期": "2026-01-01", "金额": 100, "摘要": "", "对方户名": ""},
            {"日期": None, "金额": None, "摘要": None, "对方户名": None},
            {"日期": None, "金额": 100, "摘要": "本期合计", "对方户名": None},
            {"日期": "日期", "金额": "金额", "摘要": "摘要", "对方户名": "对方户名"},
        ]
    )
    raw_journal = pd.DataFrame(
        [{"日期": "2026-01-01", "金额": 100, "摘要": "测试", "对方户名": "甲"}]
    )
    report = _build_report(
        raw_bank,
        raw_journal,
        _standardized([("2026-01-01", 100, "")], "bank"),
        _standardized([("2026-01-01", 100, "测试")], "journal"),
        parse_errors=[
            {
                "type": "日期解析失败",
                "source_type": "bank",
                "row": 5,
                "original_value": "日期错误",
                "column": "日期",
            }
        ],
    )

    assert report.has_blockers is False
    assert next(item for item in report.items if item.name == "非交易行").status == "提示"
    assert next(item for item in report.items if item.name == "辅助文字完整性").status == "提示"
    assert "少量日期或金额解析失败" in report.warning_message()


def test_日期或金额全部无可用记录会阻止():
    raw = pd.DataFrame([{"日期": "错误", "金额": "错误", "摘要": "测试", "对方户名": "甲"}])
    empty = pd.DataFrame(columns=["date", "amount", "summary", "source"])

    report = _build_report(raw, raw, empty, empty)

    assert report.has_blockers is True
    assert next(item for item in report.items if item.name == "日期范围").status == "阻止"
    assert next(item for item in report.items if item.name == "金额合计").status == "阻止"


def test_检查结果可直接转换为固定列报告表():
    raw = pd.DataFrame(
        [{"日期": "2026-01-01", "金额": 100, "摘要": "测试", "对方户名": "甲"}]
    )
    report = _build_report(
        raw,
        raw,
        _standardized([("2026-01-01", 100, "测试")], "bank"),
        _standardized([("2026-01-01", 100, "测试")], "journal"),
    )

    table = report.to_dataframe()

    assert list(table.columns) == [
        "检查项目",
        "银行流水结果",
        "银行日记账结果",
        "双方比较结果",
        "状态",
        "说明",
    ]
    assert table["检查项目"].tolist() == [
        "文件读取",
        "表格结构",
        "日期范围",
        "金额方向",
        "金额合计",
        "非交易行",
        "必填字段",
        "辅助文字完整性",
    ]


def _write_direction_pair(
    tmp_path,
    *,
    bank_amount=100,
    journal_amount=100,
    bank_direction="贷",
    journal_direction="借",
):
    bank_path = tmp_path / "银行流水.xlsx"
    journal_path = tmp_path / "银行日记账.xlsx"
    pd.DataFrame(
        [
            {
                "日期": "2026-01-01",
                "金额": bank_amount,
                "方向": bank_direction,
                "摘要": "测试收款",
            }
        ]
    ).to_excel(bank_path, index=False)
    pd.DataFrame(
        [
            {
                "日期": "2026-01-01",
                "金额": journal_amount,
                "方向": journal_direction,
                "摘要": "测试收款",
            }
        ]
    ).to_excel(journal_path, index=False)
    mapping = {
        "date": "日期",
        "amount": "金额",
        "direction": "方向",
        "summary": "摘要",
        "auxiliary_text_columns": ["摘要"],
        "mode": "single_amount_with_direction",
    }
    return bank_path, journal_path, mapping


def test_无界面入口遇到无效方向会阻止且不创建报告(tmp_path):
    bank_path, journal_path, mapping = _write_direction_pair(
        tmp_path,
        bank_direction="未知方向",
    )
    output_path = tmp_path / "不应生成.xlsx"

    with pytest.raises(input_precheck.InputPrecheckBlockedError) as exc_info:
        run_reconciliation(
            bank_path=str(bank_path),
            journal_path=str(journal_path),
            bank_mapping=mapping,
            journal_mapping=mapping,
            matcher_config=MatcherConfig(),
            output_path=output_path,
        )

    assert "方向" in str(exc_info.value)
    assert not output_path.exists()


def test_普通提示回调可以返回调整且不创建报告(tmp_path):
    bank_path, journal_path, mapping = _write_direction_pair(
        tmp_path,
        bank_amount=100,
        journal_amount=90,
    )
    output_path = tmp_path / "用户已返回调整.xlsx"
    received = []

    with pytest.raises(InterruptedError, match="返回调整"):
        run_reconciliation(
            bank_path=str(bank_path),
            journal_path=str(journal_path),
            bank_mapping=mapping,
            journal_mapping=mapping,
            matcher_config=MatcherConfig(),
            output_path=output_path,
            precheck_warning_callback=lambda report: received.append(report) or False,
        )

    assert len(received) == 1
    assert received[0].has_warnings is True
    assert not output_path.exists()


def test_最终报告包含与本次运行一致的输入预检查工作表(tmp_path):
    bank_path, journal_path, mapping = _write_direction_pair(tmp_path)
    output_path = tmp_path / "核对报告.xlsx"

    result = run_reconciliation(
        bank_path=str(bank_path),
        journal_path=str(journal_path),
        bank_mapping=mapping,
        journal_mapping=mapping,
        matcher_config=MatcherConfig(),
        output_path=output_path,
    )

    assert result == output_path
    assert "输入预检查" in pd.ExcelFile(output_path).sheet_names
    table = pd.read_excel(output_path, sheet_name="输入预检查")
    business_columns = [
        column for column in table.columns if not str(column).startswith("Unnamed:")
    ]
    assert business_columns == [
        "检查项目",
        "银行流水结果",
        "银行日记账结果",
        "双方比较结果",
        "状态",
        "说明",
    ]
    assert table.loc[table["检查项目"] == "金额合计", "状态"].item() == "通过"


class _Variable:
    def __init__(self, value=""):
        self.value = value

    def get(self):
        return self.value

    def set(self, value):
        self.value = value


def test_gui_自动检测同时采用表头位置层级和派生列名(tmp_path):
    path = _build_merged_header_workbook(tmp_path)
    loaded = []

    class _FakeApp:
        bank_path = _Variable(str(path))
        journal_path = _Variable()
        bank_skip = _Variable("0")
        journal_skip = _Variable("0")
        bank_header_rows = _Variable("1")
        journal_header_rows = _Variable("1")
        loader = DataLoader()

        def log(self, _message):
            return None

        def load_columns(self, *args):
            loaded.append(args)

    app = _FakeApp()
    gui.ReconciliationApp.auto_detect(app, "bank")

    assert app.bank_skip.get() == "1"
    assert app.bank_header_rows.get() == "2"
    assert loaded == [
        (
            str(path),
            "bank",
            1,
            2,
            [
                "日期",
                "发生额｜借方",
                "发生额｜贷方",
                "辅助信息｜摘要",
                "辅助信息｜对方户名",
            ],
        )
    ]


def test_gui_运行时把表头层级和普通提示回调交给统一入口(monkeypatch, tmp_path):
    captured = {}

    def fake_run_reconciliation(**kwargs):
        captured.update(kwargs)
        return tmp_path / "报告.xlsx"

    monkeypatch.setattr(gui, "run_reconciliation", fake_run_reconciliation)

    class _FakeApp:
        matcher = None

        def log(self, _message):
            return None

        def _set_progress(self, _value):
            return None

        def _set_stop_enabled(self, _enabled):
            return None

        def _set_start_enabled(self, _enabled):
            return None

        def _confirm_precheck_warnings(self, _report):
            return True

    state = {
        "bank_path": "银行.xlsx",
        "journal_path": "日记账.xlsx",
        "bank_mapping": {},
        "journal_mapping": {},
        "config": MatcherConfig(),
        "llm_config": None,
        "bank_skip": 1,
        "journal_skip": 2,
        "bank_header_rows": 2,
        "journal_header_rows": 1,
        "date_format": "auto",
    }

    app = _FakeApp()
    gui.ReconciliationApp.run_process(app, state)

    assert captured["bank_header_rows"] == 2
    assert captured["journal_header_rows"] == 1
    callback = captured["precheck_warning_callback"]
    assert callback.__self__ is app


def test_gui_普通提示在主线程集中确认(monkeypatch):
    raw = pd.DataFrame(
        [{"日期": "2026-01-01", "金额": 100, "摘要": "", "对方户名": ""}]
    )
    report = _build_report(
        raw,
        raw,
        _standardized([("2026-01-01", 100, "")], "bank"),
        _standardized([("2026-01-01", 90, "")], "journal"),
    )
    shown = []

    def fake_askokcancel(title, message, **_kwargs):
        shown.append((title, message))
        return True

    monkeypatch.setattr(gui.messagebox, "askokcancel", fake_askokcancel)

    class _FakeApp:
        def after(self, _delay, callback):
            callback()

    decision = gui.ReconciliationApp._confirm_precheck_warnings(
        _FakeApp(),
        report,
    )

    assert decision is True
    assert shown[0][0] == "输入预检查提示"
    assert "继续核对" in shown[0][1]
    assert "返回调整" in shown[0][1]
